from app_settings.utils import log_action
import os
import shlex
import subprocess
import time
import csv
import io
import json
import zipfile
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from urllib.parse import quote_plus, urlparse
from urllib.request import urlopen
from urllib import request as urllib_request, error as urllib_error
from xml.sax.saxutils import escape
import re

from django.contrib import messages
from django.conf import settings
from django.db import connection
from django.db.utils import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q
from django.db.models import Count
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from recruitment_teams.models import RecruitmentTeamMaster

from app_settings.access_control import has_action_permission
from app_settings.models import (
    AssessmentField,
    AssessmentForm,
    City,
    CompanyInfo,
    Country,
    EducationLevel,
    EmailDeliveryConfig,
    EmailPlaceholder,
    EmailTemplate,
    GoogleOAuthToken,
    InterviewIntegrationSetting,
    RoleMaster,
    RolePermissionSetup,
    State,
    UiNotification,
    UserMaster,
    UserMasterAudit,
    TaskSlaPrioritySetting,
)
from onboarding.models import OnboardingFormTemplate
from onboarding.utils import log_onboarding_action
from candidate_management.models import Candidate
from job_requisition.models import JobRequisition

try:
    from google_auth_oauthlib.flow import Flow
except Exception:
    Flow = None

CALENDAR_SCOPES = ["https://www.googleapis.com/auth/calendar"]

VIDEO_INTERVIEW_PROVIDERS = [
    ("google_meet", "Google Meet"),
    ("microsoft_teams", "Microsoft Teams"),
    ("zoom", "Zoom"),
]


def _extract_first_json_object(raw_text: str) -> dict:
    if not raw_text:
        return {}
    text = str(raw_text)
    start = text.find("{")
    if start < 0:
        return {}
    depth = 0
    for idx in range(start, len(text)):
        ch = text[idx]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                snippet = text[start : idx + 1].strip()
                try:
                    return json.loads(snippet)
                except Exception:
                    return {}
    return {}


def _gemini_generate_assessment_questions(
    *,
    title: str,
    description: str,
    department: str,
    difficulty_level: str,
    count: int = 20,
) -> tuple[bool, str, list[dict]]:
    api_key = (getattr(settings, "GEMINI_ASSESSMENT_API_KEY", "") or "").strip()
    model = (getattr(settings, "GEMINI_MODEL", "gemini-2.0-flash") or "gemini-2.0-flash").strip()
    if not api_key:
        return False, "Gemini API key is missing. Set GEMINI_ASSESSMENT_API_KEY in environment.", []

    count = int(count or 20)
    count = max(1, min(count, 50))

    prompt = (
        "You are an assessment question generator. Return valid JSON only (no markdown).\n"
        "Generate exactly the requested number of multiple-choice questions.\n"
        "Each question must have 4 options and exactly one correct answer.\n"
        "Keep questions aligned to the department and difficulty.\n\n"
        "Return JSON in this schema:\n"
        "{"
        '"questions": ['
        "{"
        '"label": \"...\",'
        '"options": [\"A\", \"B\", \"C\", \"D\"],'
        '"correct_answer": \"one of the options exactly\",'
        '"points\": 1'
        "}"
        "]}"
        "\n\n"
        f"Assessment title: {title}\n"
        f"Assessment description: {description or 'Not specified'}\n"
        f"Department: {department or 'General'}\n"
        f"Difficulty level: {difficulty_level or 'Medium'}\n"
        f"Number of questions: {count}\n"
    )

    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.4,
            "maxOutputTokens": 2400,
            "responseMimeType": "application/json",
        },
    }

    fallback_models = [
        model,
        "gemini-2.0-flash",
        "gemini-2.0-flash-lite",
        "gemini-1.5-flash",
    ]
    seen = set()
    ordered_models = []
    for model_name in fallback_models:
        if model_name and model_name not in seen:
            ordered_models.append(model_name)
            seen.add(model_name)

    body = None
    errors = []
    for model_name in ordered_models:
        endpoint = (
            f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent"
            f"?key={api_key}"
        )
        try:
            req = urllib_request.Request(
                endpoint,
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib_request.urlopen(req, timeout=25) as resp:
                body = json.loads(resp.read().decode("utf-8"))
            break
        except urllib_error.HTTPError as exc:
            detail = ""
            try:
                detail = exc.read().decode("utf-8", errors="ignore")
            except Exception:
                detail = str(exc)
            errors.append(f"{model_name}: {detail[:280]}")
            continue
        except Exception as exc:
            errors.append(f"{model_name}: {str(exc)}")
            continue

    if not body:
        summary = "; ".join(errors[-3:]) if errors else "Unknown error"
        return False, f"Gemini request failed. Tried models: {', '.join(ordered_models)}. Errors: {summary}", []

    text_parts = []
    for candidate in body.get("candidates", []):
        content = candidate.get("content") or {}
        for part in content.get("parts", []):
            if part.get("text"):
                text_parts.append(part["text"])
    raw_text = "\n".join(text_parts).strip()
    structured = _extract_first_json_object(raw_text)
    questions = structured.get("questions") if isinstance(structured, dict) else None
    if not isinstance(questions, list) or not questions:
        return False, "Could not parse Gemini response into questions.", []

    cleaned = []
    for item in questions:
        if not isinstance(item, dict):
            continue
        label = (item.get("label") or "").strip()
        options = item.get("options")
        correct = (item.get("correct_answer") or "").strip()
        points = item.get("points", 1)
        if not label or not isinstance(options, list) or len(options) < 2:
            continue
        options_clean = [str(opt).strip() for opt in options if str(opt).strip()]
        if len(options_clean) < 2:
            continue
        if correct and correct not in options_clean:
            correct = ""
        try:
            points_val = float(points) if points not in ("", None) else 1.0
        except Exception:
            points_val = 1.0
        cleaned.append(
            {
                "label": label[:120],
                "options": options_clean[:12],
                "correct_answer": correct[:255],
                "points": points_val,
            }
        )

    if not cleaned:
        return False, "Gemini returned no usable questions.", []

    return True, "", cleaned[:count]


def _ollama_generate_assessment_questions(
    *,
    title: str,
    description: str,
    department: str,
    difficulty_level: str,
    count: int = 20,
) -> tuple[bool, str, list[dict]]:
    if not bool(getattr(settings, "OLLAMA_ENABLED", False)):
        return False, "Ollama is disabled. Set OLLAMA_ENABLED=1 in environment.", []

    count = int(count or 20)
    count = max(1, min(count, 50))

    prompt = (
        "You are an assessment question generator.\n"
        "Return valid JSON only (no markdown, no extra text).\n"
        "Generate exactly the requested number of multiple-choice questions.\n"
        "Each question must have 4 options and exactly one correct answer.\n"
        "Keep questions aligned to the department and difficulty.\n\n"
        "Return JSON in this schema:\n"
        "{"
        '"questions": ['
        "{"
        '"label": \"...\",'
        '"options": [\"A\", \"B\", \"C\", \"D\"],'
        '"correct_answer": \"one of the options exactly\",'
        '"points\": 1'
        "}"
        "]}"
        "\n\n"
        f"Assessment title: {title}\n"
        f"Assessment description: {description or 'Not specified'}\n"
        f"Department: {department or 'General'}\n"
        f"Difficulty level: {difficulty_level or 'Medium'}\n"
        f"Number of questions: {count}\n"
    )

    try:
        from recruitment_tracking_system.ollama_client import ollama_chat
    except Exception as exc:
        return False, f"Could not import Ollama client: {exc}", []

    result = ollama_chat(
        base_url=getattr(settings, "OLLAMA_BASE_URL", "http://127.0.0.1:11434"),
        model=getattr(settings, "OLLAMA_MODEL", "llama3.1"),
        user_message=prompt,
        system_message="You output JSON only.",
        timeout_seconds=int(getattr(settings, "OLLAMA_ASSESSMENT_TIMEOUT_SECONDS", getattr(settings, "OLLAMA_TIMEOUT_SECONDS", 60)) or 60),
    )
    if not result.ok:
        err = result.error or "Ollama request failed."
        if "Read timed out" in err or "timed out" in err.lower():
            err = (
                f"{err} Increase OLLAMA_ASSESSMENT_TIMEOUT_SECONDS (e.g. 180/300) or generate fewer questions."
            )
        return False, err, []

    structured = _extract_first_json_object(result.message)
    questions = structured.get("questions") if isinstance(structured, dict) else None
    if not isinstance(questions, list) or not questions:
        return False, "Could not parse Ollama response into questions.", []

    cleaned = []
    for item in questions:
        if not isinstance(item, dict):
            continue
        label = (item.get("label") or "").strip()
        options = item.get("options")
        correct = (item.get("correct_answer") or "").strip()
        points = item.get("points", 1)
        if not label or not isinstance(options, list) or len(options) < 2:
            continue
        options_clean = [str(opt).strip() for opt in options if str(opt).strip()]
        if len(options_clean) < 2:
            continue
        if correct and correct not in options_clean:
            correct = ""
        try:
            points_val = float(points) if points not in ("", None) else 1.0
        except Exception:
            points_val = 1.0
        cleaned.append(
            {
                "label": label[:120],
                "options": options_clean[:12],
                "correct_answer": correct[:255],
                "points": points_val,
            }
        )

    if not cleaned:
        return False, "Ollama returned no usable questions.", []

    return True, "", cleaned[:count]


def _generate_assessment_questions(
    *,
    title: str,
    description: str,
    department: str,
    difficulty_level: str,
    count: int = 20,
) -> tuple[bool, str, list[dict]]:
    provider = (getattr(settings, "ASSESSMENT_LLM_PROVIDER", "") or "").strip().lower() or "ollama"
    if provider == "gemini":
        return _gemini_generate_assessment_questions(
            title=title,
            description=description,
            department=department,
            difficulty_level=difficulty_level,
            count=count,
        )

    # default: ollama
    return _ollama_generate_assessment_questions(
        title=title,
        description=description,
        department=department,
        difficulty_level=difficulty_level,
        count=count,
    )


def global_search_view(request):
    """
    Global search endpoint used by the topbar search.

    - If query matches a module keyword (e.g. "evaluation", "jobs") -> redirect to that module.
    - If query is an exact Job ID (e.g. MPR0001) -> redirect to job posting form (edit mode).
    - If query is an exact Candidate ID (e.g. CAND0001) -> redirect to candidate profile.
    - Otherwise render a results page with matching jobs/candidates.
    """

    query = (request.GET.get("q") or "").strip()
    if not query:
        return redirect("/dashboard/")

    query_l = query.strip().lower()

    shortcuts = [
        ({"dashboard", "home"}, "/dashboard/"),
        ({"candidate", "candidates", "candidate management", "cv", "resume"}, "/candidate-management/"),
        ({"job", "jobs", "job requisition", "requisition", "mpr"}, "/job-requisition/"),
        ({"pipeline", "application pipeline", "applications", "applicant tracking"}, "/applicant-tracking/application-pipeline/"),
        ({"interview", "interviews", "interview scheduling", "schedule"}, "/applicant-tracking/interview-scheduling/"),
        ({"evaluation", "eval", "interview evaluation"}, "/interview-evaluation/"),
        ({"recording", "interview recording"}, "/interview-recording/"),
        ({"bgv", "background verification", "background check"}, "/background-verification/"),
        ({"onboarding"}, "/onboarding/board/"),
        ({"task", "tasks", "task management"}, "/task-management/"),
        ({"team", "teams", "recruitment teams"}, "/recruitment-teams/"),
        ({"audit", "audit logs", "logs"}, "/settings/audit-logs/"),
        ({"onboarding audit", "onboarding audit logs"}, "/settings/onboarding-audit-logs/"),
        ({"masters", "master", "master data"}, "/settings/masters/"),
        ({"members", "member", "users"}, "/settings/members/"),
        ({"email", "email template", "email templates"}, "/settings/email-templates/"),
        ({"feedback", "feedback form"}, "/settings/feedback-form/"),
        ({"bulk", "bulk upload", "upload"}, "/settings/bulk-upload/"),
        ({"integration", "integrations", "google calendar", "calendar"}, "/settings/integration/"),
        ({"workflow"}, "/settings/workflow/"),
        ({"data export", "export"}, "/settings/data-export/"),
        ({"subscription", "billing"}, "/settings/subscription-billing/"),
        ({"company", "company info", "careers config"}, "/settings/company/"),
        ({"role", "roles"}, "/settings/role/"),
        ({"permission", "permissions", "permission management"}, "/settings/permission-management/"),
        ({"settings", "configuration"}, "/settings/"),
    ]

    # Module shortcut redirects first (so "evaluation" goes straight to that screen).
    for keywords, path in shortcuts:
        if any(kw in query_l for kw in keywords):
            return redirect(path)

    token = re.sub(r"\s+", "", query).upper()

    # Local imports to avoid circular dependencies at import time.
    from candidate_management.models import Candidate
    from job_requisition.models import JobRequisition

    if re.fullmatch(r"MPR\d{4,}", token):
        job = JobRequisition.objects.filter(job_id__iexact=token).only("job_id").first()
        if job:
            return redirect(f"{reverse('job_requisition:posting_form')}?job_id={job.job_id}")

    if re.fullmatch(r"CAND\d{4,}", token):
        candidate = Candidate.objects.filter(candidate_id__iexact=token).only("candidate_id").first()
        if candidate:
            return redirect(f"/candidate-management/profile/?candidate_id={candidate.candidate_id}")

    candidate_qs = Candidate.objects.filter(
        Q(candidate_id__icontains=query) | Q(full_name__icontains=query) | Q(email__icontains=query)
    ).order_by("-created_at")

    candidate_count = candidate_qs.count()
    if candidate_count == 1:
        candidate = candidate_qs.only("candidate_id").first()
        return redirect(f"/candidate-management/profile/?candidate_id={candidate.candidate_id}")
    if candidate_count > 1:
        return redirect(f"/candidate-management/list/?q={quote_plus(query)}")

    job_qs = JobRequisition.objects.filter(
        Q(job_id__icontains=query) | Q(title__icontains=query) | Q(department__icontains=query)
    ).order_by("-created_at")

    job_count = job_qs.count()
    if job_count == 1:
        job = job_qs.only("job_id").first()
        return redirect(f"{reverse('job_requisition:posting_form')}?job_id={job.job_id}")
    if job_count > 1:
        return redirect("/job-requisition/job-position-dashboard/")

    messages.info(request, f'No results found for "{query}".')
    return redirect("/dashboard/")

def _get_current_user(request):
    email = (request.session.get("login_user_email") or "").strip()
    return UserMaster.objects.filter(email_id__iexact=email).first()


def _normalize_smtp_password(host: str, password: str) -> str:
    cleaned_host = (host or "").strip().lower()
    cleaned_password = (password or "").strip()
    if not cleaned_password:
        return ""

    if cleaned_host in {"smtp.gmail.com", "smtp.googlemail.com"}:
        if " " in cleaned_password:
            compact = cleaned_password.replace(" ", "")
            if len(compact) == 16:
                return compact
    return cleaned_password

MASTER_CONFIG = {
    "city": {
        "model": City,
        "title": "City Form",
        "button": "Add City",
        "name_label": "City Name",
        "code_label": "City Code",
        "empty": "No cities added.",
    },
    "state": {
        "model": State,
        "title": "State Form",
        "button": "Add State",
        "name_label": "State Name",
        "code_label": "State Code",
        "empty": "No states added.",
    },
    "country": {
        "model": Country,
        "title": "Country Form",
        "button": "Add Country",
        "name_label": "Country Name",
        "code_label": "Country Code",
        "empty": "No countries added.",
    },
    "education-level": {
        "model": EducationLevel,
        "title": "Education Level Form",
        "button": "Add Education Level",
        "name_label": "Education Level Name",
        "code_label": "Education Level Code",
        "empty": "No education levels added.",
    },
}

USER_MODULE_OPTIONS = [
    {"value": "dashboard", "label": "Dashboard"},
    {"value": "candidate_management", "label": "Candidate Management"},
    {"value": "job_requisition", "label": "Requisition Management"},
    {"value": "applicant_tracking", "label": "Applicant Tracking"},
    {"value": "background_verification", "label": "Background Verification"},
    {"value": "interview_recording", "label": "Interview Recording"},
    {"value": "interview_evaluation", "label": "Interview Evaluation"},
    {"value": "candidate_portal", "label": "Candidate Portal"},
    {"value": "recruitment_teams", "label": "Recruitment Teams"},
    {"value": "task_management", "label": "Task Management"},
    {"value": "candidate_database", "label": "Candidate Database"},
    {"value": "onboarding", "label": "Onboarding"},
    {"value": "settings", "label": "Settings"},
]

PERMISSION_SUBMODULE_MAP = {
    "Dashboard": ["Overview", "Analytics", "Calendar", "Charts"],
    "Candidate Management": ["Registration", "Candidate List", "Profile", "Evaluation"],
    "Requisition Management": ["Job Posting Form", "Position Dashboard", "Approval Workflow", "Public Job URL"],
    "Applicant Tracking": ["Application Pipeline", "Interview Scheduling", "In Person Interview", "Video Interview"],
    "Background Verification": ["Submission", "Verification Dashboard", "Report", "Alerts"],
    "Interview Recording": ["Recording Dashboard", "Panel Dashboard"],
    "Interview Evaluation": ["Interview Panel", "Evaluation", "Aggregates"],
    "Candidate Portal": ["Public Jobs", "Candidate Applications", "Profile"],
    "Recruitment Teams": ["Team Dashboard", "Team Mapping", "Task Assign"],
    "Task Management": ["Task Board", "Task Assignment", "Task Status"],
    "Candidate Database": ["Candidate Records", "Filters", "Export"],
    "Onboarding": ["Board", "Offers", "Approvals", "Employee Code", "Reports"],
    "Settings": ["Assessment Form", "Masters", "User Master", "Permission Management", "Integration"],
}

MODULE_SINGLE_EXPORT_TABLE = {
    "candidate_management": ("candidate_management_candidate", "Candidate Management"),
    "candidate_database": ("candidate_database_candidatedatabasenote", "Candidate Database"),
    "job_requisition": ("job_requisition_jobrequisition", "Requisition Management"),
    "applicant_tracking": ("applicant_tracking_applicationpipelinecandidate", "Applicant Tracking"),
    "background_verification": ("background_verification_backgroundverificationrecord", "Background Verification"),
    "interview_recording": ("interview_recording_interviewrecording", "Interview Recording"),
    "interview_evaluation": ("interview_evaluation_interviewevaluationsubmission", "Interview Evaluation"),
    "task_management": ("task_management_taskrecord", "Task Management"),
    "recruitment_teams": ("recruitment_teams_recruitmentteammaster", "Recruitment Teams"),
    "candidate_portal": ("candidate_portal_candidateportalprofile", "Candidate Portal"),
    "onboarding": ("onboarding_offerletter", "Onboarding"),
}
TABLE_MODULE_LABEL_BY_NAME = {table_name: module_label for table_name, module_label in MODULE_SINGLE_EXPORT_TABLE.values()}
MODULE_BY_PATH_PREFIX = [
    ("/candidate-management/", "candidate_management"),
    ("/candidate-database/", "candidate_database"),
    ("/job-requisition/", "job_requisition"),
    ("/applicant-tracking/", "applicant_tracking"),
    ("/background-verification/", "background_verification"),
    ("/interview-recording/", "interview_recording"),
    ("/interview-evaluation/", "interview_evaluation"),
    ("/task-management/", "task_management"),
    ("/recruitment-teams/", "recruitment_teams"),
    ("/candidate-portal/", "candidate_portal"),
    ("/onboarding/", "onboarding"),
    ("/settings/", "settings"),
]

FIELD_TYPE_DISPLAY = {
    "text": "Text",
    "number": "Number",
    "date": "Date",
    "email": "Email",
    "tel": "Phone",
    "url": "URL",
    "time": "Time",
    "datetime": "Date & Time",
    "textarea": "Long Text",
    "select": "Dropdown",
    "radio": "Radio",
    "checkbox": "Checkbox",
}

def _table_name(form_id):
    return f"assessment_form_{form_id}"


def _column_sql_type(field_type):
    return {
        "number": "REAL",
        "date": "TEXT",
        "email": "TEXT",
        "tel": "TEXT",
        "url": "TEXT",
        "time": "TEXT",
        "datetime": "TEXT",
        "textarea": "TEXT",
        "select": "TEXT",
        "radio": "TEXT",
        "checkbox": "TEXT",
        "text": "TEXT",
    }.get(field_type, "TEXT")


def _existing_columns(table_name):
    with connection.cursor() as cursor:
        description = connection.introspection.get_table_description(cursor, table_name)
    return {col.name for col in description}


def _sync_dynamic_table(form):
    table = _table_name(form.id)
    quote = connection.ops.quote_name
    table_quoted = quote(table)

    with connection.cursor() as cursor:
        existing_tables = set(connection.introspection.table_names(cursor))
        if table not in existing_tables:
            cursor.execute(
                f"CREATE TABLE {table_quoted} (id INTEGER PRIMARY KEY AUTOINCREMENT, created_at TEXT)"
            )

        existing_cols = _existing_columns(table)
        for field in form.fields.all():
            col = field.field_name
            if col in existing_cols:
                continue
            null_sql = "NOT NULL" if field.required else ""
            col_type = _column_sql_type(field.field_type)
            cursor.execute(
                f"ALTER TABLE {table_quoted} ADD COLUMN {quote(col)} {col_type} {null_sql}"
            )


def _drop_dynamic_table(form_id):
    table = _table_name(form_id)
    quote = connection.ops.quote_name
    table_quoted = quote(table)
    with connection.cursor() as cursor:
        existing_tables = set(connection.introspection.table_names(cursor))
        if table in existing_tables:
            cursor.execute(f"DROP TABLE {table_quoted}")


def app_settings_view(request):
    forms = AssessmentForm.objects.order_by("-created_at")
    return render(
        request,
        "app_settings/index.html",
        {"forms": forms, "settings_menu_active": "settings_home"},
    )


def assessment_form_builder_view(request, form_id=None):
    # Assessment Builder is deprecated. Always redirect to Assessment Wizard.
    if form_id:
        return redirect(f"{reverse('app_settings:assessment_wizard', kwargs={'form_id': form_id})}?step=1")
    return redirect("app_settings:assessment_wizard")

    forms = list(AssessmentForm.objects.order_by("-created_at"))
    selected_form = None
    if form_id:
        selected_form = get_object_or_404(AssessmentForm, pk=form_id)
    elif request.GET.get("form_id"):
        raw_form_id = request.GET["form_id"]
        selected_form = get_object_or_404(AssessmentForm, pk=raw_form_id)
    elif forms:
        selected_form = forms[0]

    selected_form_fields = list(selected_form.fields.all()) if selected_form else []

    return render(
        request,
        "app_settings/assessment_builder.html",
        {
            "forms": forms,
            "selected_form": selected_form,
            "selected_form_fields": selected_form_fields,
            "settings_menu_active": "assessment_form",
        },
    )


def assessment_form_wizard_view(request, form_id=None):
    """
    Multi-step assessment creation flow:
      1) Basic Details
      2) Questions
      3) Additional
      4) Review
    """
    if not form_id and request.method == "GET" and not (request.GET.get("step") or "").strip():
        forms = list(
            AssessmentForm.objects.annotate(fields_count=Count("fields"))
            .order_by("-created_at")
            .values(
                "id",
                "name",
                "form_type",
                "department",
                "difficulty_level",
                "time_limit_minutes",
                "passing_score_percent",
                "is_published",
                "created_at",
                "fields_count",
            )
        )
        return render(
            request,
            "app_settings/assessment_wizard_list.html",
            {
                "forms": forms,
                "settings_menu_active": "assessment_form",
            },
        )

    step = 1
    try:
        step = int((request.GET.get("step") or "1").strip() or 1)
    except Exception:
        step = 1
    step = max(1, min(step, 4))

    form_obj = None
    if form_id:
        form_obj = get_object_or_404(AssessmentForm, pk=form_id)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        def _wizard_url(target_form_id: int | None, target_step: int) -> str:
            if target_form_id:
                base = reverse("app_settings:assessment_wizard", kwargs={"form_id": int(target_form_id)})
            else:
                base = reverse("app_settings:assessment_wizard")
            return f"{base}?step={int(target_step)}"

        if action == "save_basic":
            name = (request.POST.get("name") or "").strip()
            description = (request.POST.get("description") or "").strip()
            department = (request.POST.get("department") or "").strip()
            difficulty_level = (request.POST.get("difficulty_level") or "").strip()
            time_limit = (request.POST.get("time_limit_minutes") or "").strip()
            passing = (request.POST.get("passing_score_percent") or "").strip()
            form_type = (request.POST.get("form_type") or "assessment").strip() or "assessment"

            if not name:
                messages.error(request, "Assessment title is required.")
                return redirect(_wizard_url(form_obj.id if form_obj else None, 1))

            time_limit_val = None
            passing_val = None
            try:
                time_limit_val = int(time_limit) if time_limit else None
            except Exception:
                time_limit_val = None
            try:
                passing_val = int(passing) if passing else None
            except Exception:
                passing_val = None

            existing_same_name = AssessmentForm.objects.filter(name__iexact=name)
            if form_obj:
                existing_same_name = existing_same_name.exclude(id=form_obj.id)
            if existing_same_name.exists():
                messages.error(request, "Assessment title already exists. Please use a different title.")
                return redirect(_wizard_url(form_obj.id if form_obj else None, 1))

            try:
                if not form_obj:
                    form_obj = AssessmentForm.objects.create(
                        name=name,
                        form_type=form_type if form_type in {"assessment", "feedback"} else "assessment",
                        description=description,
                        department=department,
                        difficulty_level=difficulty_level if difficulty_level in {"Easy", "Medium", "Hard"} else "",
                        time_limit_minutes=time_limit_val,
                        passing_score_percent=passing_val,
                    )
                    _sync_dynamic_table(form_obj)
                else:
                    form_obj.name = name
                    form_obj.form_type = form_type if form_type in {"assessment", "feedback"} else form_obj.form_type
                    form_obj.description = description
                    form_obj.department = department
                    form_obj.difficulty_level = difficulty_level if difficulty_level in {"Easy", "Medium", "Hard"} else ""
                    form_obj.time_limit_minutes = time_limit_val
                    form_obj.passing_score_percent = passing_val
                    form_obj.save(
                        update_fields=[
                            "name",
                            "form_type",
                            "description",
                            "department",
                            "difficulty_level",
                            "time_limit_minutes",
                            "passing_score_percent",
                        ]
                    )
            except IntegrityError:
                messages.error(request, "Assessment title already exists. Please use a different title.")
                return redirect(_wizard_url(form_obj.id if form_obj else None, 1))

            messages.success(request, "Basic details saved.")
            return redirect(_wizard_url(form_obj.id, 2))

        if not form_obj:
            messages.error(request, "Please create assessment basic details first.")
            return redirect(_wizard_url(None, 1))

        if action in {"add_field", "update_field", "delete_field"}:
            if action == "add_field":
                label = request.POST.get("label", "").strip()
                field_type = request.POST.get("field_type", "text")
                required = request.POST.get("required") == "on"
                options = request.POST.get("options", "").strip()
                points_raw = (request.POST.get("points") or "").strip()
                correct_answer = (request.POST.get("correct_answer") or "").strip()
                raw_field_name = request.POST.get("field_name", "").strip() or label
                field_name = slugify(raw_field_name).replace("-", "_")

                if not label or not field_name:
                    messages.error(request, "Question label is required.")
                    return redirect(_wizard_url(form_obj.id, 2))

                points = 1.0
                try:
                    points = float(points_raw) if points_raw else 1.0
                except Exception:
                    points = 1.0

                AssessmentField.objects.create(
                    form=form_obj,
                    label=label,
                    field_name=field_name,
                    field_type=field_type,
                    required=required,
                    options=options,
                    points=points,
                    correct_answer=correct_answer,
                )
                _sync_dynamic_table(form_obj)
                messages.success(request, "Question added.")
                return redirect(_wizard_url(form_obj.id, 2))

            if action == "update_field":
                raw_field_id = request.POST.get("field_id")
                field = get_object_or_404(AssessmentField, pk=raw_field_id, form=form_obj)
                field.label = (request.POST.get("label") or field.label).strip()[:120] or field.label
                field.required = request.POST.get("required") == "on"
                field.options = (request.POST.get("options") or "").strip()
                points_raw = (request.POST.get("points") or "").strip()
                correct_answer = (request.POST.get("correct_answer") or "").strip()
                try:
                    field.points = float(points_raw) if points_raw else field.points
                except Exception:
                    pass
                field.correct_answer = correct_answer
                field.save(update_fields=["label", "required", "options", "points", "correct_answer"])
                _sync_dynamic_table(form_obj)
                messages.success(request, "Question updated.")
                return redirect(_wizard_url(form_obj.id, 2))

            if action == "delete_field":
                raw_field_id = request.POST.get("field_id")
                field = get_object_or_404(AssessmentField, pk=raw_field_id, form=form_obj)
                field.delete()
                messages.success(request, "Question removed.")
                return redirect(_wizard_url(form_obj.id, 2))

        if action == "generate_ai_questions":
            if step != 2:
                return redirect(_wizard_url(form_obj.id, 2))

            desired_count = (request.POST.get("questions_count") or "20").strip()
            try:
                desired_count_int = int(desired_count)
            except Exception:
                desired_count_int = 20
            desired_count_int = max(1, min(desired_count_int, 50))

            ok, err, questions = _generate_assessment_questions(
                title=form_obj.name,
                description=form_obj.description,
                department=form_obj.department,
                difficulty_level=form_obj.difficulty_level,
                count=desired_count_int,
            )
            if not ok:
                messages.error(request, err or "Failed to generate questions.")
                return redirect(_wizard_url(form_obj.id, 2))

            existing_names = set(form_obj.fields.values_list("field_name", flat=True))

            def _unique_field_name(base: str) -> str:
                base = slugify(base or "question").replace("-", "_")[:60] or "question"
                candidate = base
                suffix = 2
                while candidate in existing_names:
                    candidate = f"{base}_{suffix}"
                    suffix += 1
                existing_names.add(candidate)
                return candidate

            rows = []
            for idx, q in enumerate(questions, start=1):
                label = (q.get("label") or "").strip()[:120]
                options = q.get("options") or []
                options_text = ", ".join([str(opt).strip() for opt in options if str(opt).strip()])
                correct_answer = (q.get("correct_answer") or "").strip()
                points = q.get("points") or 1.0

                field_name_seed = f"q_{idx}"
                field_name = _unique_field_name(field_name_seed)
                rows.append(
                    AssessmentField(
                        form=form_obj,
                        label=label,
                        field_name=field_name,
                        field_type="radio",
                        required=True,
                        options=options_text,
                        points=float(points or 1.0),
                        correct_answer=correct_answer,
                    )
                )

            if rows:
                AssessmentField.objects.bulk_create(rows)
                _sync_dynamic_table(form_obj)

            messages.success(request, f"Generated {len(rows)} questions.")
            return redirect(_wizard_url(form_obj.id, 2))

        if action == "save_additional":
            form_obj.instructions = (request.POST.get("instructions") or "").strip()
            form_obj.shuffle_questions = request.POST.get("shuffle_questions") == "on"
            form_obj.show_score_to_candidate = request.POST.get("show_score_to_candidate") == "on"
            form_obj.allow_multiple_attempts = request.POST.get("allow_multiple_attempts") == "on"
            form_obj.save(
                update_fields=[
                    "instructions",
                    "shuffle_questions",
                    "show_score_to_candidate",
                    "allow_multiple_attempts",
                ]
            )
            messages.success(request, "Additional settings saved.")
            return redirect(_wizard_url(form_obj.id, 4))

        if action == "publish":
            form_obj.is_published = True
            form_obj.save(update_fields=["is_published"])
            messages.success(request, "Assessment published.")
            return redirect(_wizard_url(form_obj.id, 4))

    selected_form_fields = list(form_obj.fields.all()) if form_obj else []

    department_options = sorted(
        {
            val.strip()
            for val in (
                list(JobRequisition.objects.exclude(department__exact="").values_list("department", flat=True))
                + list(UserMaster.objects.exclude(department__exact="").values_list("department", flat=True))
                + list(AssessmentForm.objects.exclude(department__exact="").values_list("department", flat=True))
            )
            if (val or "").strip()
        }
    )

    return render(
        request,
        "app_settings/assessment_wizard.html",
        {
            "step": step,
            "selected_form": form_obj,
            "selected_form_fields": selected_form_fields,
            "department_options": department_options,
            "settings_menu_active": "assessment_form",
        },
    )


def assessment_form_delete_confirm_view(request, form_id):
    form_obj = get_object_or_404(AssessmentForm, pk=form_id)
    if request.method == "POST":
        if request.POST.get("action", "").strip() == "confirm_delete":
            _drop_dynamic_table(form_obj.id)
            form_obj.delete()
            messages.success(request, "Form deleted.")
        return redirect("app_settings:assessment_wizard")
    return render(
        request,
        "app_settings/assessment_delete_confirm.html",
        {"form_obj": form_obj, "settings_menu_active": "assessment_form"},
    )


def assessment_form_fill_view(request, form_id):
    form = get_object_or_404(AssessmentForm, pk=form_id)
    _sync_dynamic_table(form)
    fields = list(form.fields.all())
    for field in fields:
        field.options_list = [item.strip() for item in field.options.split(",") if item.strip()]

    if request.method == "POST":
        if not fields:
            table = _table_name(form.id)
            quote = connection.ops.quote_name
            with connection.cursor() as cursor:
                cursor.execute(
                    f"INSERT INTO {quote(table)} ({quote('created_at')}) VALUES (%s)",
                    [timezone.now().isoformat(timespec="seconds")],
                )
            messages.success(request, "Assessment entry saved.")
            return redirect("app_settings:assessment_fill", form_id=form.id)

        table = _table_name(form.id)
        quote = connection.ops.quote_name
        cols = []
        values = []
        placeholders = []

        for field in fields:
            if field.field_type == "checkbox":
                values_list = [item.strip() for item in request.POST.getlist(field.field_name) if item.strip()]
                value = ", ".join(values_list)
            else:
                value = request.POST.get(field.field_name, "").strip()
            if field.required and not value:
                messages.error(request, f"{field.label} is required.")
                return redirect("app_settings:assessment_fill", form_id=form.id)
            cols.append(quote(field.field_name))
            values.append(value)
            placeholders.append("%s")

        with connection.cursor() as cursor:
            cursor.execute(
                f"INSERT INTO {quote(table)} ({', '.join(cols)}) VALUES ({', '.join(placeholders)})",
                values,
            )
        messages.success(request, "Assessment entry saved.")
        return redirect("app_settings:assessment_fill", form_id=form.id)

    demo_entries = []

    entry_columns = [field.label for field in fields]
    entry_rows = []
    for item in demo_entries:
        entry_rows.append([item.get(field.field_name, "") for field in fields])

    return render(
        request,
        "app_settings/assessment_fill.html",
        {
            "form_obj": form,
            "fields": fields,
            "entry_columns": entry_columns,
            "entry_rows": entry_rows,
            "settings_menu_active": "assessment_form",
        },
    )


def _settings_module_view(request, module_key, module_title):
    demo_rows_map = {
        "members": [
            "Aman Sharma (Admin)",
            "Neha Verma (Recruiter)",
            "Rohit Singh (Interviewer)",
        ],
        "email_templates": [
            "Interview Invite Template",
            "Offer Letter Template",
            "BGV Pending Documents Reminder",
        ],
        "feedback_form": [
            "Technical Interview Feedback",
            "HR Round Feedback",
        ],
        "bulk_upload": [
            "Candidate Bulk Upload - Last run: 2026-02-22",
            "Job Posting Bulk Upload - Last run: 2026-02-20",
        ],
        "integration": [
            "Email Gateway - Connected",
            "Calendar Sync - Connected",
            "Document Storage - Connected",
        ],
        "data_export": [
            "Candidate Report Export (CSV)",
            "Interview Scores Export (Excel)",
        ],
        "subscription_billing": [
            "Plan: Enterprise ATS",
            "Next Renewal: 2026-03-15",
        ],
    }
    return render(
        request,
        "app_settings/module_placeholder.html",
        {
            "settings_menu_active": module_key,
            "module_title": module_title,
            "module_rows": demo_rows_map.get(module_key, []),
        },
    )


def masters_view(request):
    return render(
        request,
        "app_settings/masters.html",
        {
            "settings_menu_active": "masters",
            "master_options": MASTER_CONFIG,
        },
    )


def masters_form_view(request, master_key):
    config = MASTER_CONFIG.get(master_key)
    if not config:
        return redirect("app_settings:masters")

    model_cls = config["model"]
    if request.method == "POST":
        action = request.POST.get("action", "create").strip()
        record_id = request.POST.get("record_id", "").strip()
        name = request.POST.get("name", "").strip()
        code = request.POST.get("code", "").strip()
        if action == "delete" and record_id:
            record = model_cls.objects.filter(id=record_id).first()
            if record:
                name_val = getattr(record, "name", str(record_id))
                record.delete()
                log_action(
                    module="app_settings",
                    action=f"MASTER_{master_key.upper().replace('-','_')}_DELETED",
                    user=_get_current_user(request),
                    details={"name": name_val},
                    request=request
                )
                messages.success(request, "Record deleted.")
            return redirect("app_settings:masters_form", master_key=master_key)

        if not name or not code:
            messages.error(request, "Name and Code are required.")
            return redirect("app_settings:masters_form", master_key=master_key)
        try:
            if action == "update" and record_id:
                record = model_cls.objects.filter(id=record_id).first()
                if record:
                    record.name = name
                    record.code = code
                    record.save()
                    log_action(
                        module="app_settings",
                        action=f"MASTER_{master_key.upper().replace('-','_')}_UPDATED",
                        user=_get_current_user(request),
                        details={"name": name},
                        request=request
                    )
                    messages.success(request, "Record updated.")
                else:
                    messages.error(request, "Record not found.")
            else:
                model_cls.objects.create(name=name, code=code)
                log_action(
                    module="app_settings",
                    action=f"MASTER_{master_key.upper().replace('-','_')}_CREATED",
                    user=_get_current_user(request),
                    details={"name": name},
                    request=request
                )
                messages.success(request, f"{config['button']} successful.")
        except IntegrityError:
            messages.error(request, "Name or Code already exists.")
        return redirect("app_settings:masters_form", master_key=master_key)

    edit_id = request.GET.get("edit", "").strip()
    selected_record = model_cls.objects.filter(id=edit_id).first() if edit_id else None
    records = list(model_cls.objects.all())

    return render(
        request,
        "app_settings/masters_form.html",
        {
            "settings_menu_active": "masters",
            "master_key": master_key,
            "master_config": config,
            "records": records,
            "selected_record": selected_record,
        },
    )


def members_view(request):
    members = list(
        UserMaster.objects.order_by("full_name", "email_id").values(
            "full_name",
            "user_id",
            "employee_code",
            "email_id",
            "mobile_number",
            "role",
            "status",
        )
    )
    return render(
        request,
        "app_settings/members.html",
        {
            "settings_menu_active": "members",
            "members": members,
        },
    )


def email_templates_view(request):
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "delete_template":
            template_id = request.POST.get("template_id", "").strip()
            template = get_object_or_404(EmailTemplate, id=template_id)
            template.delete()
            
            # Log template deletion
            current_email = (request.session.get("login_user_email") or "").strip()
            user = UserMaster.objects.filter(email_id__iexact=current_email).first()
            log_onboarding_action(
                user=user,
                action="TEMPLATE_DELETED",
                details={
                    "template_id": template_id,
                    "source": "Settings > Email Templates"
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            
            messages.success(request, "Email template deleted.")
            return redirect("app_settings:email_templates")

        template_id = request.POST.get("template_id", "").strip()
        name = request.POST.get("name", "").strip()
        module = request.POST.get("module", "").strip()
        trigger = request.POST.get("trigger", "").strip()
        to_emails = request.POST.get("to_emails", "").strip()
        cc_emails = request.POST.get("cc_emails", "").strip()
        bcc_emails = request.POST.get("bcc_emails", "").strip()
        subject = request.POST.get("subject", "").strip()
        body = request.POST.get("body", "").strip()
        is_active = request.POST.get("is_active") == "on"
        onboarding_fields = request.POST.getlist("onboarding_fields")
        onboarding_documents = request.POST.getlist("onboarding_documents")
        if not name or not subject or not body:
            messages.error(request, "Name, subject, and body are required.")
        else:
            template_obj = None
            if template_id:
                template_obj = get_object_or_404(EmailTemplate, id=template_id)
                template_obj.name = name
                template_obj.module = module
                template_obj.trigger = trigger
                template_obj.to_emails = to_emails
                template_obj.cc_emails = cc_emails
                template_obj.bcc_emails = bcc_emails
                template_obj.subject = subject
                template_obj.body = body
                template_obj.is_active = is_active
                template_obj.save()
                messages.success(request, "Email template updated.")
            else:
                template_obj = EmailTemplate.objects.create(
                    name=name,
                    module=module,
                    trigger=trigger,
                    to_emails=to_emails,
                    cc_emails=cc_emails,
                    bcc_emails=bcc_emails,
                    subject=subject,
                    body=body,
                    is_active=is_active,
                )
                messages.success(request, "Email template created.")
            # Sync onboarding form template (optional)
            if template_obj:
                normalized_module = (template_obj.module or "").strip().lower()
                normalized_trigger = (template_obj.trigger or "").strip().lower()
                is_onboarding_offer = (
                    "onboarding" in normalized_module
                    or "offer" in normalized_trigger
                    or "offer" in (template_obj.name or "").lower()
                )
                if is_onboarding_offer and (onboarding_fields or onboarding_documents):
                    OnboardingFormTemplate.objects.update_or_create(
                        offer_email_template=template_obj,
                        defaults={
                            "name": f"{template_obj.name} - Onboarding",
                            "description": "Managed via Settings > Email Templates",
                            "selected_fields": onboarding_fields,
                            "document_requirements": onboarding_documents,
                            "is_active": template_obj.is_active,
                            "created_by": "Settings Email Templates",
                        },
                    )
                    
                    # Log template update
                    current_email = (request.session.get("login_user_email") or "").strip()
                    user = UserMaster.objects.filter(email_id__iexact=current_email).first()
                    log_onboarding_action(
                        user=user,
                        action="TEMPLATE_UPDATED",
                        details={
                            "template_name": template_obj.name,
                            "source": "Settings > Email Templates"
                        },
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT')
                    )
            return redirect("app_settings:email_templates")

    edit_id = request.GET.get("edit_id", "").strip()
    create_mode = (request.GET.get("create") or "").strip().lower() in {"1", "true", "yes"}
    edit_template = EmailTemplate.objects.filter(id=edit_id).first() if edit_id else None
    onboarding_config = None
    if edit_template:
        onboarding_config = OnboardingFormTemplate.objects.filter(offer_email_template=edit_template).first()
    show_form = bool(edit_template) or create_mode
    templates = EmailTemplate.objects.order_by("-updated_at")
    placeholders = EmailPlaceholder.objects.order_by("key")
    return render(
        request,
        "app_settings/email_templates.html",
        {
            "settings_menu_active": "email_templates",
            "templates": templates,
            "edit_template": edit_template,
            "create_mode": create_mode,
            "show_form": show_form,
            "onboarding_config": onboarding_config,
            "onboarding_field_options": [
                ("full_name", "Full Name"),
                ("email", "Email"),
                ("contact_number", "Contact Number"),
                ("date_of_birth", "Date of Birth"),
                ("gender", "Gender"),
                ("address", "Address"),
                ("current_city", "Current City"),
                ("state", "State"),
                ("country", "Country"),
                ("experience", "Previous Experience (years)"),
                ("employment_history", "Employment History"),
                ("skills", "Skills"),
                ("pan", "PAN"),
                ("aadhaar", "Aadhaar"),
                ("pf", "PF Number"),
                ("esic", "ESIC Number"),
            ],
            "onboarding_document_options": [
                ("pan", "PAN Card"),
                ("aadhaar", "Aadhaar Card"),
                ("resume", "Resume"),
                ("photo", "Passport-size Photo"),
                ("address_proof", "Address Proof"),
                ("education_certificate", "Education Certificate"),
                ("experience_letter", "Experience Letter"),
            ],
            "placeholders": placeholders,
        },
    )


def email_template_preview_view(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Invalid request method."}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    subject = (payload.get("subject") or "").strip()
    body = (payload.get("body") or "").strip()
    module = (payload.get("module") or "").strip()
    trigger = (payload.get("trigger") or "").strip()

    candidate = Candidate.objects.order_by("-updated_at").first()
    job = JobRequisition.objects.order_by("-created_at").first()
    today = timezone.localdate()

    sample = {
        "company_name": getattr(settings, "ATS_COMPANY_NAME", "Ultimatix ATS"),
        "candidate_name": getattr(candidate, "full_name", "") or "Candidate",
        "candidate_email": getattr(candidate, "email", "") or "candidate@example.com",
        "candidate_id": getattr(candidate, "candidate_id", "") or "CAND0001",
        "job_title": getattr(job, "title", "") or "Job Title",
        "job_id": getattr(job, "job_id", "") or "MPR0001",
        "previous_stage": "Applied",
        "new_stage": "Interview",
        "stage": "Interview",
        "portal_url": "/candidate-portal/",
        "today": today.isoformat(),
        "module": module,
        "trigger": trigger,
    }

    def _render(text: str) -> str:
        rendered = text or ""
        for key, value in sample.items():
            rendered = rendered.replace("{{" + key + "}}", str(value or ""))
        return rendered

    rendered_subject = _render(subject)
    rendered_body = _render(body)
    if rendered_body and "<html" not in rendered_body.lower():
        rendered_body = f"<div style=\"font-family: Arial, sans-serif; font-size: 14px; line-height: 1.6;\">{rendered_body}</div>"

    return JsonResponse(
        {
            "ok": True,
            "rendered_subject": rendered_subject,
            "rendered_body": rendered_body,
            "sample": sample,
        }
    )


def feedback_form_view(request):
    return _settings_module_view(request, "feedback_form", "Feedback Form")


def _has_bulk_upload_permission(request, action_key: str) -> bool:
    """
    Bulk upload is used to import candidates. Operationally, this is a Candidate Management
    action even though the screen lives under Settings.

    Allow access if the logged-in role has either Settings permission OR Candidate Management
    permission for the requested action.
    """

    role = request.session.get("login_user_role")
    return has_action_permission(role, "settings", action_key) or has_action_permission(
        role, "candidate_management", action_key
    )


def bulk_upload_view(request):
    if not _has_bulk_upload_permission(request, "create"):
        messages.error(request, "You do not have permission to access Bulk Upload.")
        return redirect("app_settings:index")

    if (request.GET.get("clear_result") or "").strip() in {"1", "true", "yes"}:
        request.session.pop("bulk_upload_result", None)
        messages.info(request, "Bulk upload result cleared.")
        return redirect("app_settings:bulk_upload")

    if request.method == "POST":
        upload = request.FILES.get("upload_file")
        entity = (request.POST.get("entity") or "candidate").strip().lower()
        upsert_mode = (request.POST.get("upsert_mode") or "create_only").strip().lower()

        if entity != "candidate":
            messages.error(request, "Unsupported bulk upload type.")
            return redirect("app_settings:bulk_upload")

        if not upload:
            messages.error(request, "Please choose a file to upload.")
            return redirect("app_settings:bulk_upload")

        ext = os.path.splitext(upload.name or "")[1].lower()
        if ext not in {".xlsx", ".csv"}:
            messages.error(request, "Unsupported file type. Please upload .xlsx or .csv")
            return redirect("app_settings:bulk_upload")

        os.makedirs(os.path.join(settings.MEDIA_ROOT, "bulk_upload", "tmp"), exist_ok=True)
        upload_id = f"{int(time.time())}_{slugify(os.path.splitext(upload.name)[0]) or 'upload'}"
        saved_name = f"{upload_id}{ext}"
        saved_path = os.path.join(settings.MEDIA_ROOT, "bulk_upload", "tmp", saved_name)

        with open(saved_path, "wb") as handle:
            for chunk in upload.chunks():
                handle.write(chunk)

        request.session["bulk_upload_ctx"] = {
            "entity": entity,
            "upsert_mode": upsert_mode,
            "upload_id": upload_id,
            "filename": upload.name,
            "path": saved_path,
            "ext": ext,
            "created_at": timezone.now().isoformat(),
        }
        return redirect("app_settings:bulk_upload_map")

    # Keep the last result in session so users can come back and review it.
    result = request.session.get("bulk_upload_result")
    last_ctx = request.session.get("bulk_upload_ctx")

    return render(
        request,
        "app_settings/bulk_upload.html",
        {
            "settings_menu_active": "bulk_upload",
            "result": result,
            "last_ctx": last_ctx,
        },
    )


def _bulk_upload_normalize_header(value):
    label = (value or "").strip()
    label = re.sub(r"[\s\-/]+", "_", label)
    label = re.sub(r"[^\w_]+", "", label)
    return label.strip("_").lower()


def _bulk_upload_candidate_fields():
    return [
        ("", "-- Skip --"),
        ("full_name", "Full Name (required)"),
        ("email", "Email (required, unique)"),
        ("contact_number", "Contact Number (required)"),
        ("date_of_birth", "Date of Birth"),
        ("gender", "Gender"),
        ("current_city", "Current City"),
        ("country", "Country"),
        ("state", "State"),
        ("pan", "PAN"),
        ("aadhaar", "Aadhaar"),
        ("pf", "PF"),
        ("esic", "ESIC"),
        ("social_media_link", "Social Media Link"),
        ("address", "Address"),
        ("resume_path", "Resume (Path/URL)"),
        ("highest_education_level", "Highest Education Level"),
        ("degree_name", "Degree Name"),
        ("institute_name", "Institute Name"),
        ("year_of_passing", "Year Of Passing"),
        ("percentage_cgpa", "Percentage/CGPA"),
        ("certifications", "Certifications"),
        ("skills", "Skills"),
        ("experience", "Experience"),
        ("employment_history", "Employment History"),
        ("references", "References"),
        ("custom_tags", "Custom Tags"),
        ("applied_position", "Applied Position"),
        ("status", "Status"),
    ]


def _bulk_upload_guess_mapping(headers, available_fields):
    field_keys = {key for key, _label in available_fields if key}
    alias_map = {
        "name": "full_name",
        "full_name": "full_name",
        "candidate_name": "full_name",
        "candidate_full_name": "full_name",
        "employee_name": "full_name",
        "applicant_name": "full_name",
        "email": "email",
        "email_id": "email",
        "emailid": "email",
        "e_mail": "email",
        "e_mail_id": "email",
        "email_address": "email",
        "candidate_email": "email",
        "mail": "email",
        "phone": "contact_number",
        "mobile": "contact_number",
        "contact": "contact_number",
        "contact_number": "contact_number",
        "contact_no": "contact_number",
        "contact_number1": "contact_number",
        "phone_number": "contact_number",
        "phone_no": "contact_number",
        "mobile_number": "contact_number",
        "mobile_no": "contact_number",
        "dob": "date_of_birth",
        "date_of_birth": "date_of_birth",
        "birth_date": "date_of_birth",
        "gender": "gender",
        "city": "current_city",
        "current_city": "current_city",
        "current_location": "current_city",
        "location": "current_city",
        "country": "country",
        "state": "state",
        "pan": "pan",
        "aadhaar": "aadhaar",
        "aadhar": "aadhaar",
        "aadhar_no": "aadhaar",
        "aadhaar_no": "aadhaar",
        "pf": "pf",
        "esic": "esic",
        "linkedin": "social_media_link",
        "linkedin_url": "social_media_link",
        "linkedin_profile": "social_media_link",
        "social_media_link": "social_media_link",
        "address": "address",
        "current_address": "address",
        "resume": "resume_path",
        "resume_url": "resume_path",
        "resume_link": "resume_path",
        "cv": "resume_path",
        "cv_url": "resume_path",
        "cv_link": "resume_path",
        "education": "highest_education_level",
        "highest_education_level": "highest_education_level",
        "highest_qualification": "highest_education_level",
        "qualification": "highest_education_level",
        "degree": "degree_name",
        "degree_name": "degree_name",
        "course": "degree_name",
        "institute": "institute_name",
        "institute_name": "institute_name",
        "college": "institute_name",
        "year_of_passing": "year_of_passing",
        "yop": "year_of_passing",
        "passing_year": "year_of_passing",
        "graduation_year": "year_of_passing",
        "cgpa": "percentage_cgpa",
        "percentage": "percentage_cgpa",
        "score": "percentage_cgpa",
        "skills": "skills",
        "experience": "experience",
        "total_experience": "experience",
        "total_exp": "experience",
        "exp": "experience",
        "years_of_experience": "experience",
        "employment_history": "employment_history",
        "references": "references",
        "position": "applied_position",
        "applied_position": "applied_position",
        "designation": "applied_position",
        "job_title": "applied_position",
        "status": "status",
    }
    mapping = {}
    used = set()
    for raw in headers:
        normalized = _bulk_upload_normalize_header(raw)
        mapped = alias_map.get(normalized, "")
        if mapped and mapped in field_keys and mapped not in used:
            mapping[raw] = mapped
            used.add(mapped)
        else:
            mapping[raw] = ""
    return mapping


def _bulk_upload_read_csv_headers_and_preview(path, preview_rows=5):
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        headers = next(reader, [])
        headers = [str(h or "").strip() for h in headers if str(h or "").strip()]
        preview = []
        for _i in range(preview_rows):
            row = next(reader, None)
            if row is None:
                break
            preview.append(row)
        return headers, preview


def _bulk_upload_read_xlsx_headers_and_preview(path, preview_rows=5):
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required to import .xlsx files") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # Some Excel sheets include a title row / blank rows before the actual header.
    # Heuristic: choose the row (within first 20) with the most non-empty cells.
    best_row_index = 1
    best_row_values = ()
    best_non_empty = 0
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_row_index = idx
            best_row_values = row
    header_row = best_row_values or next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(h or "").strip() for h in header_row if str(h or "").strip()]

    preview = []
    start_row = best_row_index + 1
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + preview_rows - 1, values_only=True):
        preview.append([row[i] if i < len(row) else "" for i in range(len(header_row))])
    return headers, preview, best_row_index, len(header_row)


def bulk_upload_map_view(request):
    if not _has_bulk_upload_permission(request, "create"):
        messages.error(request, "You do not have permission to access Bulk Upload.")
        return redirect("app_settings:index")

    ctx = request.session.get("bulk_upload_ctx") or {}
    if not ctx.get("path") or not os.path.exists(ctx["path"]):
        messages.error(request, "Upload session expired. Please upload the file again.")
        return redirect("app_settings:bulk_upload")

    available_fields = _bulk_upload_candidate_fields()
    saved_mapping = request.session.get("bulk_upload_mapping") or {}

    try:
        if ctx.get("ext") == ".csv":
            headers, preview = _bulk_upload_read_csv_headers_and_preview(ctx["path"])
        else:
            headers, preview, header_row_index, header_row_len = _bulk_upload_read_xlsx_headers_and_preview(ctx["path"])
            ctx["xlsx_header_row_index"] = header_row_index
            ctx["xlsx_header_row_len"] = header_row_len
            request.session["bulk_upload_ctx"] = ctx
    except Exception as exc:
        messages.error(request, f"Unable to read file: {exc}")
        return redirect("app_settings:bulk_upload")

    guessed = _bulk_upload_guess_mapping(headers, available_fields)
    mapping_rows = [
        {
            "idx": idx,
            "header": header,
            "guessed": (saved_mapping.get(header) or guessed.get(header, "")),
        }
        for idx, header in enumerate(headers)
    ]

    if request.method == "POST":
        mapping = {}
        chosen = set()
        for idx, header in enumerate(headers):
            key = request.POST.get(f"map__{idx}", "").strip()
            if key:
                if key in chosen:
                    messages.error(request, "Each field can only be mapped once. Please fix duplicates.")
                    mapping[header] = key
                    continue
                chosen.add(key)
            mapping[header] = key

        request.session["bulk_upload_mapping"] = mapping

        # If duplicates were detected, re-render and keep the user's selections instead of redirecting
        # back to auto-guessed values.
        if len(chosen) != len([v for v in mapping.values() if v]):
            mapping_rows = [
                {"idx": idx, "header": header, "guessed": mapping.get(header, "")}
                for idx, header in enumerate(headers)
            ]
            return render(
                request,
                "app_settings/bulk_upload_map.html",
                {
                    "settings_menu_active": "bulk_upload",
                    "ctx": ctx,
                    "headers": headers,
                    "preview_rows": preview,
                    "available_fields": available_fields,
                    "mapping_rows": mapping_rows,
                },
            )

        return redirect("app_settings:bulk_upload_run")

    return render(
        request,
        "app_settings/bulk_upload_map.html",
        {
            "settings_menu_active": "bulk_upload",
            "ctx": ctx,
            "headers": headers,
            "preview_rows": preview,
            "available_fields": available_fields,
            "mapping_rows": mapping_rows,
        },
    )


def _bulk_upload_parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None

    # Handle HR-style strings like: "32 y (29 May 1991)" or "30y(28 May 1993)"
    # Extract the date part inside parentheses when present.
    m = re.search(r"\(([^)]+)\)", text)
    if m:
        text = (m.group(1) or "").strip()
    # Also handle strings like: "32 y 29 May 1991"
    text = re.sub(r"^\s*\d+\s*(y|yr|yrs|year|years)\s*", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\s+", " ", text).strip()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass

    # Common Excel-export formats
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def _bulk_upload_clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _bulk_upload_is_valid_resume_path(value: str) -> bool:
    """
    Bulk upload can only store a resume reference (URL or a media-relative PDF path)
    in Candidate.resume_path. Free-text summaries must not be mapped to resume_path.
    """

    text = (value or "").strip()
    if not text:
        return False
    if re.search(r"[\r\n]", text):
        return False
    if len(text) > 255:
        return False

    try:
        parsed = urlparse(text)
        if parsed.scheme in {"http", "https"}:
            return True
    except Exception:
        pass

    lower = text.lower()
    if lower.endswith(".pdf"):
        return True

    return False


def bulk_upload_run_view(request):
    if not _has_bulk_upload_permission(request, "create"):
        messages.error(request, "You do not have permission to access Bulk Upload.")
        return redirect("app_settings:index")

    ctx = request.session.get("bulk_upload_ctx") or {}
    mapping = request.session.get("bulk_upload_mapping") or {}
    if not ctx.get("path") or not os.path.exists(ctx["path"]) or not mapping:
        messages.error(request, "Upload session expired. Please upload the file again.")
        return redirect("app_settings:bulk_upload")

    if request.method != "POST":
        return render(
            request,
            "app_settings/bulk_upload_run.html",
            {
                "settings_menu_active": "bulk_upload",
                "ctx": ctx,
                "mapping": mapping,
            },
        )

    from candidate_management.models import Candidate

    required = {"full_name", "email", "contact_number"}
    mapped_fields = {v for v in mapping.values() if v}
    if not required.issubset(mapped_fields):
        missing = ", ".join(sorted(required - mapped_fields))
        messages.error(request, f"Missing required mappings: {missing}")
        return redirect("app_settings:bulk_upload_map")

    upsert_mode = (ctx.get("upsert_mode") or "create_only").strip().lower()
    allow_update = upsert_mode == "update_existing"

    is_csv = ctx.get("ext") == ".csv"
    headers = []
    xlsx_header_row_index = int(ctx.get("xlsx_header_row_index") or 1)
    try:
        if not is_csv:
            try:
                import openpyxl
            except Exception as exc:
                raise RuntimeError("openpyxl is required to import .xlsx files") from exc
            wb = openpyxl.load_workbook(ctx["path"], read_only=True, data_only=True)
            ws = wb.worksheets[0]
            header_row = next(ws.iter_rows(min_row=xlsx_header_row_index, max_row=xlsx_header_row_index, values_only=True), ())
            headers = [str(h or "").strip() for h in header_row]
    except Exception as exc:
        messages.error(request, f"Unable to read file: {exc}")
        return redirect("app_settings:bulk_upload")

    successes = 0
    errors = []
    total_rows = 0
    imported_rows = []

    def import_row(row_index, row):
        nonlocal successes, total_rows, imported_rows
        total_rows += 1

        candidate_data = {}
        for header, field_name in mapping.items():
            if not field_name:
                continue
            raw_value = row.get(header, "")
            if field_name == "date_of_birth":
                parsed = _bulk_upload_parse_date(raw_value)
                if raw_value not in ("", None) and not parsed:
                    errors.append({"row": row_index, "error": f"Invalid date_of_birth: {raw_value}"})
                    return
                candidate_data[field_name] = parsed
                continue

            cleaned = _bulk_upload_clean_cell(raw_value)
            if cleaned == "" and field_name == "status":
                continue
            if cleaned and field_name == "resume_path":
                if not _bulk_upload_is_valid_resume_path(cleaned):
                    errors.append(
                        {
                            "row": row_index,
                            "error": "Invalid resume value. Map resume_path only to a PDF URL/path (e.g. https://.../resume.pdf).",
                        }
                    )
                    continue
            candidate_data[field_name] = cleaned

        email = (candidate_data.get("email") or "").strip().lower()
        if not email:
            errors.append({"row": row_index, "error": "Email is required"})
            return
        candidate_data["email"] = email

        if not (candidate_data.get("full_name") or "").strip():
            errors.append({"row": row_index, "error": "Full Name is required"})
            return
        if not (candidate_data.get("contact_number") or "").strip():
            errors.append({"row": row_index, "error": "Contact Number is required"})
            return

        # Portal Attribution: candidates imported via Bulk Upload should be marked as Imported
        # (this value is shown in Candidate Management list under "Portal Attribution").
        if not (candidate_data.get("references") or "").strip():
            candidate_data["references"] = "Imported"

        try:
            existing = Candidate.objects.filter(email__iexact=email).first()
            if existing:
                if not allow_update:
                    errors.append({"row": row_index, "error": "Duplicate email (already exists)"})
                    return
                for key, value in candidate_data.items():
                    setattr(existing, key, value)
                if not (existing.references or "").strip():
                    existing.references = "Imported"
                existing.save()
                successes += 1
                imported_rows.append(
                    {
                        "row": row_index,
                        "action": "updated",
                        "candidate_id": existing.candidate_id,
                        "full_name": existing.full_name,
                        "email": existing.email,
                    }
                )
                return

            created = Candidate.objects.create(**candidate_data)
            successes += 1
            imported_rows.append(
                {
                    "row": row_index,
                    "action": "created",
                    "candidate_id": created.candidate_id,
                    "full_name": created.full_name,
                    "email": created.email,
                }
            )
        except IntegrityError as exc:
            errors.append({"row": row_index, "error": f"Integrity error: {exc}"})
        except Exception as exc:
            errors.append({"row": row_index, "error": str(exc)})

    try:
        if is_csv:
            with open(ctx["path"], "r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                for row_index, row in enumerate(reader, start=2):
                    if not row or not any((str(v or "").strip() for v in row.values())):
                        continue
                    import_row(row_index, row)
        else:
            start_row = xlsx_header_row_index + 1
            for row_index, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                payload = {}
                for idx, header in enumerate(headers):
                    payload[header] = row[idx] if idx < len(row) else ""
                import_row(row_index, payload)
    except Exception as exc:
        messages.error(request, f"Import failed: {exc}")
        return redirect("app_settings:bulk_upload")

    request.session["bulk_upload_result"] = {
        "successes": successes,
        "errors": errors[:5000],
        "total_rows": total_rows,
        "imported_rows": imported_rows[:200],
        "uploaded_at": timezone.now().isoformat(),
        "filename": ctx.get("filename"),
        "upsert_mode": upsert_mode,
    }

    try:
        log_action(
            request,
            action_type="bulk_upload",
            module="app_settings",
            details=json.dumps(
                {
                    "entity": ctx.get("entity"),
                    "filename": ctx.get("filename"),
                    "upload_id": ctx.get("upload_id"),
                    "successes": successes,
                    "errors": len(errors),
                    "upsert_mode": upsert_mode,
                }
            ),
        )
    except Exception:
        pass

    messages.success(request, f"Bulk upload finished. Imported/Updated: {successes}. Errors: {len(errors)}.")
    return redirect("app_settings:bulk_upload")


def bulk_upload_template_view(request):
    if not _has_bulk_upload_permission(request, "view"):
        messages.error(request, "You do not have permission to download templates.")
        return redirect("app_settings:index")

    try:
        import openpyxl
    except Exception as exc:
        messages.error(request, "openpyxl is required to generate templates.")
        return redirect("app_settings:bulk_upload")

    headers = [
        "Full Name",
        "Email",
        "Contact Number",
        "Date of Birth",
        "Gender",
        "Current City",
        "Country",
        "State",
        "PAN",
        "Aadhaar",
        "Resume URL",
        "Skills",
        "Experience",
        "Applied Position",
        "Status",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(headers)
    ws.append(
        [
            "Aman Sharma",
            "aman.sharma@example.com",
            "9999999999",
            "1997-01-25",
            "Male",
            "Pune",
            "India",
            "Maharashtra",
            "ABCDE1234F",
            "1234 5678 9012",
            "https://example.com/resume.pdf",
            "Python, Django",
            "3",
            "HR Manager",
            "Applied",
        ]
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="candidate_bulk_upload_template.xlsx"'
    return response


def integration_view(request):
    existing = {
        item.provider_key: item
        for item in InterviewIntegrationSetting.objects.filter(
            provider_key__in=[key for key, _label in VIDEO_INTERVIEW_PROVIDERS]
        )
    }
    settings_rows = []
    for key, label in VIDEO_INTERVIEW_PROVIDERS:
        row = existing.get(key)
        if not row:
            row = InterviewIntegrationSetting.objects.create(
                provider_key=key,
                provider_label=label,
                is_enabled=False,
                meeting_url="",
            )
        settings_rows.append(row)

    email_config, _created = EmailDeliveryConfig.objects.get_or_create(
        id=1,
        defaults={
            "smtp_enabled": False,
            "host": "",
            "port": 587,
            "username": "",
            "password": "",
            "use_tls": True,
            "from_email": "",
        },
    )

    if request.method == "POST":
        posted_google_oauth_fields = any(
            (request.POST.get(name, "") or "").strip()
            for name in (
                "google_client_id",
                "google_client_secret",
                "google_auth_uri",
                "google_token_uri",
                "google_cert_url",
                "google_redirect_uris",
            )
        )
        for row in settings_rows:
            row.is_enabled = request.POST.get(f"{row.provider_key}_enabled") == "on"
            row.meeting_url = request.POST.get(f"{row.provider_key}_meeting_url", "").strip()
            row.organizer_email = request.POST.get(f"{row.provider_key}_organizer_email", "").strip()
            credential_value = request.POST.get(f"{row.provider_key}_credential_json", "").strip()
            if row.provider_key == "google_meet":
                # Google OAuth credentials are loaded from environment variables
                # (see GOOGLE_OAUTH_* in `.env`). Avoid persisting client secrets to DB.
                credential_value = ""
            row.credential_json = credential_value
            row.save(update_fields=["is_enabled", "meeting_url", "organizer_email", "credential_json", "updated_at"])

        email_config.smtp_enabled = request.POST.get("smtp_enabled") == "on"
        email_config.host = request.POST.get("smtp_host", "").strip()
        email_config.port = int(request.POST.get("smtp_port", "587") or 587)
        email_config.username = request.POST.get("smtp_username", "").strip()
        email_config.password = _normalize_smtp_password(
            email_config.host,
            request.POST.get("smtp_password", ""),
        )
        email_config.use_tls = request.POST.get("smtp_use_tls") == "on"
        email_config.from_email = request.POST.get("smtp_from_email", "").strip()
        email_config.save()
        if posted_google_oauth_fields:
            messages.info(
                request,
                "Google OAuth credentials are now read from `.env` (GOOGLE_OAUTH_CLIENT_ID / GOOGLE_OAUTH_CLIENT_SECRET / GOOGLE_OAUTH_REDIRECT_URIS).",
            )
        messages.success(request, "Integration settings updated.")
        return redirect("app_settings:integration")

    google_oauth_prefill = {}
    google_row = next((item for item in settings_rows if item.provider_key == "google_meet"), None)
    env_client_id = (getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "") or "").strip()
    env_client_secret = (getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "") or "").strip()
    env_auth_uri = (getattr(settings, "GOOGLE_OAUTH_AUTH_URI", "") or "").strip() or "https://accounts.google.com/o/oauth2/auth"
    env_token_uri = (getattr(settings, "GOOGLE_OAUTH_TOKEN_URI", "") or "").strip() or "https://oauth2.googleapis.com/token"
    env_cert_url = (getattr(settings, "GOOGLE_OAUTH_CERT_URL", "") or "").strip() or "https://www.googleapis.com/oauth2/v1/certs"
    env_redirect_uris = getattr(settings, "GOOGLE_OAUTH_REDIRECT_URIS", []) or []

    if env_client_id or env_client_secret or env_redirect_uris:
        # Prefer env values; do not send the secret back to the template.
        google_oauth_prefill = {
            "client_id": env_client_id,
            "client_secret": "",
            "secret_configured": bool(env_client_secret),
            "auth_uri": env_auth_uri,
            "token_uri": env_token_uri,
            "cert_url": env_cert_url,
            "redirect_uris": ", ".join(env_redirect_uris),
            "source": "env",
        }
    elif google_row and (google_row.credential_json or "").strip():
        # Backward-compatible fallback: legacy DB-stored config.
        try:
            payload = json.loads(google_row.credential_json)
            client = payload.get("web") or payload.get("installed") or payload
            google_oauth_prefill = {
                "client_id": client.get("client_id", ""),
                "client_secret": "",
                "secret_configured": bool(client.get("client_secret", "")),
                "auth_uri": client.get("auth_uri", "https://accounts.google.com/o/oauth2/auth"),
                "token_uri": client.get("token_uri", "https://oauth2.googleapis.com/token"),
                "cert_url": client.get("auth_provider_x509_cert_url", "https://www.googleapis.com/oauth2/v1/certs"),
                "redirect_uris": ", ".join(client.get("redirect_uris", []) or []),
                "source": "db",
            }
        except Exception:
            google_oauth_prefill = {}

    current_email = (request.session.get("login_user_email") or "").strip()
    current_name = (request.session.get("login_user_name") or "").strip()
    current_role = (request.session.get("login_user_role") or "").strip().lower()
    current_user = None
    if current_email:
        current_user = UserMaster.objects.filter(email_id__iexact=current_email).first()
    if not current_user and current_name:
        current_user = UserMaster.objects.filter(full_name__iexact=current_name).first()
    session_has_login = bool(current_email or current_name)
    user_missing = session_has_login and current_user is None
    is_admin_session = current_role in {"admin", "super admin", "administrator", "dashboard user"}
    calendar_token = None
    if current_user:
        calendar_token = GoogleOAuthToken.objects.filter(user=current_user, provider="google_calendar").first()

    return render(
        request,
        "app_settings/integration.html",
        {
            "settings_menu_active": "integration",
            "provider_rows": settings_rows,
            "email_config": email_config,
            "current_user": current_user,
            "calendar_token": calendar_token,
            "google_oauth_prefill": google_oauth_prefill,
            "session_email": current_email,
            "session_name": current_name,
            "session_has_login": session_has_login,
            "user_missing": user_missing,
            "is_admin_session": is_admin_session,
        },
    )


def _get_google_oauth_client_config(setting):
    env_client_id = (getattr(settings, "GOOGLE_OAUTH_CLIENT_ID", "") or "").strip()
    env_client_secret = (getattr(settings, "GOOGLE_OAUTH_CLIENT_SECRET", "") or "").strip()
    if env_client_id and env_client_secret:
        config = {
            "web": {
                "client_id": env_client_id,
                "client_secret": env_client_secret,
                "auth_uri": getattr(
                    settings,
                    "GOOGLE_OAUTH_AUTH_URI",
                    "https://accounts.google.com/o/oauth2/auth",
                ),
                "token_uri": getattr(
                    settings,
                    "GOOGLE_OAUTH_TOKEN_URI",
                    "https://oauth2.googleapis.com/token",
                ),
                "auth_provider_x509_cert_url": getattr(
                    settings,
                    "GOOGLE_OAUTH_CERT_URL",
                    "https://www.googleapis.com/oauth2/v1/certs",
                ),
                "redirect_uris": getattr(settings, "GOOGLE_OAUTH_REDIRECT_URIS", []) or [],
            }
        }
        return config, ""

    if not setting or not (setting.credential_json or "").strip():
        return None, "Google OAuth client JSON is empty."
    try:
        payload = json.loads(setting.credential_json)
    except Exception as exc:
        return None, f"Invalid credential JSON: {exc}"

    if "installed" in payload:
        client = payload["installed"]
    elif "web" in payload:
        client = payload["web"]
    else:
        client = payload

    if not client.get("client_id") or not client.get("client_secret"):
        return None, "Credential JSON must contain client_id and client_secret."

    config = {
        "web": {
            "client_id": client.get("client_id"),
            "client_secret": client.get("client_secret"),
            "auth_uri": client.get("auth_uri", "https://accounts.google.com/o/oauth2/auth"),
            "token_uri": client.get("token_uri", "https://oauth2.googleapis.com/token"),
            "auth_provider_x509_cert_url": client.get(
                "auth_provider_x509_cert_url",
                "https://www.googleapis.com/oauth2/v1/certs",
            ),
            "redirect_uris": client.get("redirect_uris", []),
        }
    }
    return config, ""


def google_calendar_connect_view(request):
    if Flow is None:
        messages.error(request, "Google OAuth dependencies are missing.")
        return redirect("app_settings:integration")

    setting = InterviewIntegrationSetting.objects.filter(provider_key="google_meet").first()
    config, error = _get_google_oauth_client_config(setting)
    if not config:
        messages.error(request, error)
        return redirect("app_settings:integration")

    redirect_uri = request.build_absolute_uri(reverse("app_settings:google_calendar_oauth_callback"))
    flow = Flow.from_client_config(config, scopes=CALENDAR_SCOPES, redirect_uri=redirect_uri)
    auth_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    request.session["google_oauth_state"] = state
    request.session["google_oauth_started_at"] = timezone.now().isoformat()
    return redirect(auth_url)


def google_calendar_oauth_callback_view(request):
    if Flow is None:
        messages.error(request, "Google OAuth dependencies are missing.")
        return redirect("app_settings:integration")

    setting = InterviewIntegrationSetting.objects.filter(provider_key="google_meet").first()
    config, error = _get_google_oauth_client_config(setting)
    if not config:
        messages.error(request, error)
        return redirect("app_settings:integration")

    state = request.session.get("google_oauth_state")
    redirect_uri = request.build_absolute_uri(reverse("app_settings:google_calendar_oauth_callback"))
    flow = Flow.from_client_config(config, scopes=CALENDAR_SCOPES, state=state, redirect_uri=redirect_uri)
    try:
        flow.fetch_token(authorization_response=request.build_absolute_uri())
    except Exception as exc:
        messages.error(request, f"OAuth failed: {exc}")
        return redirect("app_settings:integration")

    current_email = (request.session.get("login_user_email") or "").strip()
    current_name = (request.session.get("login_user_name") or "").strip()
    user = None
    if current_email:
        user = UserMaster.objects.filter(email_id__iexact=current_email).first()
    if not user and current_name:
        user = UserMaster.objects.filter(full_name__iexact=current_name).first()

    if not user:
        # Fallback: resolve user from Google profile when session is missing.
        google_email = ""
        try:
            import requests

            profile_resp = requests.get(
                "https://www.googleapis.com/oauth2/v2/userinfo",
                headers={"Authorization": f"Bearer {flow.credentials.token}"},
                timeout=8,
            )
            if profile_resp.ok:
                google_email = (profile_resp.json().get("email") or "").strip()
        except Exception:
            google_email = ""

        if google_email:
            user = UserMaster.objects.filter(email_id__iexact=google_email).first()

    if not user:
        messages.error(request, "Logged-in user not found for OAuth save.")
        return redirect("app_settings:integration")

    creds = flow.credentials
    GoogleOAuthToken.objects.update_or_create(
        user=user,
        provider="google_calendar",
        defaults={
            "access_token": creds.token or "",
            "refresh_token": creds.refresh_token or "",
            "token_uri": creds.token_uri or "https://oauth2.googleapis.com/token",
            "scopes": " ".join(creds.scopes or []),
            "expiry": creds.expiry,
        },
    )
    messages.success(request, "Google Calendar connected for your account.")
    return redirect("app_settings:integration")


def google_calendar_disconnect_view(request):
    current_email = (request.session.get("login_user_email") or "").strip()
    current_name = (request.session.get("login_user_name") or "").strip()
    user = None
    if current_email:
        user = UserMaster.objects.filter(email_id__iexact=current_email).first()
    if not user and current_name:
        user = UserMaster.objects.filter(full_name__iexact=current_name).first()
    if not user:
        messages.error(request, "Logged-in user not found.")
        return redirect("app_settings:integration")

    GoogleOAuthToken.objects.filter(user=user, provider="google_calendar").delete()
    messages.success(request, "Google Calendar disconnected.")
    return redirect("app_settings:integration")


def workflow_view(request):
    dashboard_url = getattr(settings, "MAUTIC_DASHBOARD_URL", "http://127.0.0.1:8080/s/dashboard")
    sso_entry_url = getattr(settings, "MAUTIC_SSO_ENTRY_URL", "").strip()
    sso_next_param = getattr(settings, "MAUTIC_SSO_NEXT_PARAM", "redirect_uri").strip() or "redirect_uri"
    start_command = getattr(settings, "MAUTIC_START_COMMAND", "").strip()
    start_workdir = getattr(settings, "MAUTIC_START_WORKDIR", "")
    wait_seconds = int(getattr(settings, "MAUTIC_START_WAIT_SECONDS", 12))

    def _url_is_ready(url):
        try:
            with urlopen(url, timeout=1.5):
                return True
        except Exception:
            return False

    # If dashboard is already up, skip any start command to avoid restart churn.
    if not _url_is_ready(dashboard_url) and start_command:
        command_parts = shlex.split(start_command, posix=os.name != "nt")
        popen_kwargs = {
            "cwd": start_workdir or None,
            "stdout": subprocess.DEVNULL,
            "stderr": subprocess.DEVNULL,
            "stdin": subprocess.DEVNULL,
            "shell": False,
        }
        if os.name == "nt":
            popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS
        else:
            popen_kwargs["start_new_session"] = True

        try:
            subprocess.Popen(command_parts, **popen_kwargs)
        except OSError as exc:
            messages.warning(request, f"Could not start Mautic automatically: {exc}")
            return redirect(dashboard_url)

        # Wait for the actual dashboard URL, not just host root.
        deadline = time.monotonic() + max(wait_seconds, 0)
        while time.monotonic() < deadline:
            if _url_is_ready(dashboard_url):
                break
            time.sleep(0.8)

    if sso_entry_url:
        sso_target = sso_entry_url
        parsed = urlparse(sso_entry_url)
        joiner = "&" if parsed.query else "?"
        if f"{sso_next_param}=" not in parsed.query:
            sso_target = f"{sso_entry_url}{joiner}{sso_next_param}={quote_plus(dashboard_url)}"
        return redirect(sso_target)

    return redirect(dashboard_url)


def _sanitize_sheet_name(name, used_names):
    cleaned = "".join(ch if ch not in '[]:*?/\\' else "_" for ch in (name or "Sheet"))
    cleaned = cleaned[:31] or "Sheet"
    base = cleaned
    index = 1
    while cleaned in used_names:
        suffix = f"_{index}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(cleaned)
    return cleaned


def _excel_xml_cell(value):
    if value is None:
        return "<Cell/>"

    if isinstance(value, bool):
        data_type = "Number"
        data_value = "1" if value else "0"
    elif isinstance(value, (int, float, Decimal)):
        data_type = "Number"
        data_value = str(value)
    elif isinstance(value, (datetime, date, dt_time)):
        data_type = "String"
        data_value = value.isoformat()
    else:
        data_type = "String"
        data_value = str(value)

    return f'<Cell><Data ss:Type="{data_type}">{escape(data_value)}</Data></Cell>'


def _build_excel_xml(table_exports):
    lines = [
        '<?xml version="1.0"?>',
        '<?mso-application progid="Excel.Sheet"?>',
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel" '
        'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">',
        '<Styles>',
        '<Style ss:ID="Header"><Font ss:Bold="1"/></Style>',
        "</Styles>",
    ]

    used_names = set()
    for item in table_exports:
        sheet_name = _sanitize_sheet_name(item["table_name"], used_names)
        lines.append(f'<Worksheet ss:Name="{escape(sheet_name)}">')
        lines.append("<Table>")

        lines.append("<Row>")
        for col in item["columns"]:
            lines.append(
                f'<Cell ss:StyleID="Header"><Data ss:Type="String">{escape(col)}</Data></Cell>'
            )
        lines.append("</Row>")

        for row in item["rows"]:
            lines.append("<Row>")
            for value in row:
                lines.append(_excel_xml_cell(value))
            lines.append("</Row>")

        lines.append("</Table>")
        lines.append("</Worksheet>")

    lines.append("</Workbook>")
    return "\n".join(lines)


def _build_csv_content(columns, rows):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(["" if value is None else value for value in row])
    return buffer.getvalue()


def _pdf_escape(text):
    raw = (text or "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    return raw.encode("latin-1", "replace").decode("latin-1")


def _build_simple_pdf(table_exports):
    page_w = 842
    page_h = 595
    margin_x = 24
    margin_y = 24
    usable_w = page_w - (margin_x * 2)
    row_h = 16
    title_gap = 20
    min_col_w = 88
    max_cols_per_block = max(1, int(usable_w // min_col_w))

    pages = []
    current = {"cmds": []}
    pages.append(current)
    y = page_h - margin_y

    def new_page():
        nonlocal current, y
        current = {"cmds": []}
        pages.append(current)
        y = page_h - margin_y

    def add_text(x, y_pos, text, size=8, bold=False):
        font = "/F2" if bold else "/F1"
        current["cmds"].append(f"BT {font} {size} Tf {x:.2f} {y_pos:.2f} Td ({_pdf_escape(text)}) Tj ET")

    def add_rect(x, y_pos, w, h):
        current["cmds"].append(f"{x:.2f} {y_pos:.2f} {w:.2f} {h:.2f} re S")

    def cell_text(value, width):
        txt = "" if value is None else str(value)
        max_chars = max(1, int((width - 6) / 4.6))
        if len(txt) > max_chars:
            return txt[: max_chars - 1] + "..."
        return txt

    def draw_table_block(table_name, cols, rows, start_idx, end_idx):
        nonlocal y
        block_cols = cols[start_idx:end_idx]
        col_count = len(block_cols)
        if col_count <= 0:
            return
        col_w = usable_w / col_count

        if y < margin_y + 80:
            new_page()

        add_text(margin_x, y, f"{table_name} ({start_idx + 1}-{end_idx}/{len(cols)} columns)", 10, True)
        y -= title_gap

        # Header row
        if y < margin_y + (row_h * 2):
            new_page()
        header_y = y - row_h
        for idx, col in enumerate(block_cols):
            x = margin_x + (idx * col_w)
            add_rect(x, header_y, col_w, row_h)
            add_text(x + 3, header_y + 5, cell_text(col, col_w), 8, True)
        y = header_y

        # Data rows
        row_cursor = 0
        while row_cursor < len(rows):
            if y < margin_y + row_h:
                new_page()
                add_text(margin_x, y, f"{table_name} (cont.)", 10, True)
                y -= title_gap
                header_y = y - row_h
                for idx, col in enumerate(block_cols):
                    x = margin_x + (idx * col_w)
                    add_rect(x, header_y, col_w, row_h)
                    add_text(x + 3, header_y + 5, cell_text(col, col_w), 8, True)
                y = header_y

            row = rows[row_cursor]
            row_y = y - row_h
            for idx in range(col_count):
                x = margin_x + (idx * col_w)
                add_rect(x, row_y, col_w, row_h)
                cell_idx = start_idx + idx
                value = row[cell_idx] if cell_idx < len(row) else ""
                add_text(x + 3, row_y + 5, cell_text(value, col_w), 8, False)
            y = row_y
            row_cursor += 1

        y -= 10

    if not table_exports:
        add_text(margin_x, y - 20, "No data available.", 10, True)
    else:
        for table in table_exports:
            cols = table.get("columns", [])
            rows = table.get("rows", [])
            if not cols:
                if y < margin_y + 50:
                    new_page()
                add_text(margin_x, y - 20, f"{table['table_name']}: No columns", 10, True)
                y -= 40
                continue
            for start in range(0, len(cols), max_cols_per_block):
                end = min(start + max_cols_per_block, len(cols))
                draw_table_block(table["table_name"], cols, rows, start, end)

    objects = []

    def add_object(content):
        objects.append(content)
        return len(objects)

    catalog_id = add_object("<< /Type /Catalog /Pages 2 0 R >>")
    add_object("<< /Type /Pages /Kids [] /Count 0 >>")
    font_regular_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_bold_id = add_object("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    kids = []
    for page in pages:
        content_stream = "\n".join(page["cmds"]).encode("latin-1", "replace")
        content_id = add_object(
            f"<< /Length {len(content_stream)} >>\nstream\n{content_stream.decode('latin-1')}\nendstream"
        )
        page_id = add_object(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] "
            f"/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> "
            f"/Contents {content_id} 0 R >>"
        )
        kids.append(f"{page_id} 0 R")

    objects[1] = f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>"

    output = io.BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f"{idx} 0 obj\n{obj}\nendobj\n".encode("latin-1", "replace"))

    xref_pos = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    output.write(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        output.write(f"{off:010d} 00000 n \n".encode("latin-1"))
    output.write(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\n"
            f"startxref\n{xref_pos}\n%%EOF"
        ).encode("latin-1")
    )
    return output.getvalue()


def _build_export_response(table_exports, export_format, file_stem):
    fmt = (export_format or "excel").strip().lower()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = slugify(file_stem or "export").replace("-", "_") or "export"

    if fmt == "excel":
        xml_payload = _build_excel_xml(table_exports)
        response = HttpResponse(xml_payload, content_type="application/vnd.ms-excel; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{safe_stem}_{stamp}.xls"'
        return response

    if fmt == "csv":
        if len(table_exports) == 1:
            item = table_exports[0]
            csv_payload = _build_csv_content(item["columns"], item["rows"])
            response = HttpResponse(csv_payload, content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = f'attachment; filename="{safe_stem}_{stamp}.csv"'
            return response

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for item in table_exports:
                table_safe = slugify(item["table_name"]).replace("-", "_") or "table"
                csv_payload = _build_csv_content(item["columns"], item["rows"])
                archive.writestr(f"{table_safe}.csv", csv_payload)
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{safe_stem}_{stamp}_csv.zip"'
        return response

    if fmt == "pdf":
        pdf_payload = _build_simple_pdf(table_exports)
        response = HttpResponse(pdf_payload, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{safe_stem}_{stamp}.pdf"'
        return response

    xml_payload = _build_excel_xml(table_exports)
    response = HttpResponse(xml_payload, content_type="application/vnd.ms-excel; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{safe_stem}_{stamp}.xls"'
    return response


def _module_key_from_path(path):
    raw_path = (path or "").strip()
    for prefix, module_key in MODULE_BY_PATH_PREFIX:
        if raw_path.startswith(prefix):
            return module_key
    return ""


def data_export_view(request):
    with connection.cursor() as cursor:
        all_tables = connection.introspection.table_names(cursor)

    excluded = {"django_migrations", "sqlite_sequence"}
    table_names = sorted([name for name in all_tables if name not in excluded])

    if request.method == "POST":
        selected_tables = request.POST.getlist("tables")
        export_format = request.POST.get("export_format", "excel").strip().lower()
        if not selected_tables:
            messages.error(request, "Select at least one table to export.")
            return redirect("app_settings:data_export")

        allowed_table_set = set(table_names)
        export_tables = [name for name in selected_tables if name in allowed_table_set]
        if not export_tables:
            messages.error(request, "No valid tables selected for export.")
            return redirect("app_settings:data_export")

        table_exports = []
        with connection.cursor() as cursor:
            quote = connection.ops.quote_name
            for table_name in export_tables:
                cursor.execute(f"SELECT * FROM {quote(table_name)}")
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]
                table_exports.append(
                    {
                        "table_name": table_name,
                        "columns": columns,
                        "rows": rows,
                    }
                )

        return _build_export_response(table_exports, export_format, "ats_data_export")

    table_items = []
    with connection.cursor() as cursor:
        quote = connection.ops.quote_name
        for table_name in table_names:
            cursor.execute(f"SELECT COUNT(*) FROM {quote(table_name)}")
            row_count = cursor.fetchone()[0]
            cursor.execute(f"SELECT * FROM {quote(table_name)} LIMIT 1")
            column_count = len(cursor.description or [])
            table_items.append(
                {
                    "name": table_name,
                    "display_name": table_name.replace("_", " "),
                    "row_count": row_count,
                    "column_count": column_count,
                    "module_label": TABLE_MODULE_LABEL_BY_NAME.get(table_name, "Common"),
                    "export_url": reverse("app_settings:data_export_table", kwargs={"table_name": table_name}),
                }
            )

    return render(
        request,
        "app_settings/data_export.html",
        {
            "settings_menu_active": "data_export",
            "table_items": table_items,
        },
    )


def data_export_table_view(request, table_name):
    role_name = request.session.get("login_user_role", "Admin")
    if not has_action_permission(role_name, "settings", "export"):
        messages.error(request, "Export permission denied for your role.")
        return redirect("app_settings:data_export")

    safe_table = (table_name or "").strip()
    export_format = request.GET.get("format", "excel").strip().lower()
    if not safe_table:
        messages.error(request, "Table name is required.")
        return redirect("app_settings:data_export")

    with connection.cursor() as cursor:
        all_tables = set(connection.introspection.table_names(cursor))
        if safe_table not in all_tables:
            messages.error(request, "Selected table is not available.")
            return redirect("app_settings:data_export")
        quote = connection.ops.quote_name
        cursor.execute(f"SELECT * FROM {quote(safe_table)}")
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    return _build_export_response(
        [
            {
                "table_name": safe_table,
                "columns": columns,
                "rows": rows,
            }
        ],
        export_format,
        safe_table,
    )


def data_export_payload_view(request):
    if request.method != "POST":
        return redirect("app_settings:data_export")

    export_format = request.POST.get("export_format", "excel").strip().lower()
    source_path = request.POST.get("source_path", "").strip()
    table_name = request.POST.get("table_name", "").strip() or "table_export"
    columns_json = request.POST.get("columns_json", "[]")
    rows_json = request.POST.get("rows_json", "[]")

    module_key = _module_key_from_path(source_path)
    role_name = request.session.get("login_user_role", "Admin")
    if module_key and not has_action_permission(role_name, module_key, "export"):
        messages.error(request, "Export permission denied for your role.")
        return redirect("/dashboard/")

    try:
        columns = json.loads(columns_json)
        rows = json.loads(rows_json)
    except json.JSONDecodeError:
        messages.error(request, "Invalid table payload for export.")
        return redirect(source_path or "/dashboard/")

    clean_columns = [str(col) for col in (columns or [])]
    clean_rows = []
    for row in rows or []:
        if isinstance(row, list):
            clean_row = ["" if cell is None else str(cell) for cell in row]
            if clean_columns:
                if len(clean_row) < len(clean_columns):
                    clean_row = clean_row + ([""] * (len(clean_columns) - len(clean_row)))
                elif len(clean_row) > len(clean_columns):
                    clean_row = clean_row[: len(clean_columns)]
            clean_rows.append(clean_row)

    return _build_export_response(
        [
            {
                "table_name": table_name,
                "columns": clean_columns,
                "rows": clean_rows,
            }
        ],
        export_format,
        table_name,
    )


def subscription_billing_view(request):
    from super_admin.models import CompanySubscription
    from super_admin.utils import current_context_from_request, is_platform_superadmin

    if is_platform_superadmin(request):
        return redirect("/super-admin/dashboard/")

    ctx = current_context_from_request(request)
    company = ctx.company
    subscription = None
    if company:
        subscription = (
            CompanySubscription.objects.filter(company=company).select_related("plan").order_by("-end_date", "-created_at").first()
        )

    return render(
        request,
        "app_settings/subscription_billing.html",
        {
            "settings_menu_active": "subscription_billing",
            "company": company,
            "subscription": subscription,
        },
    )


def company_view(request):
    return render(
        request,
        "app_settings/company/index.html",
        {"settings_menu_active": "company"},
    )


def company_info_view(request):
    from super_admin.utils import current_context_from_request, is_platform_superadmin

    if is_platform_superadmin(request):
        return redirect("/super-admin/dashboard/")

    ctx = current_context_from_request(request)
    root_company = ctx.company
    if not root_company:
        messages.error(request, "No main company configured. Please contact platform Super Admin.")
        return redirect("/dashboard/")

    preview_data = None
    selected_company = None
    edit_id = (request.GET.get("edit") or "").strip()
    if edit_id:
        selected_company = CompanyInfo.objects.filter(id=edit_id).first()
        if selected_company and selected_company.get_root_company().id != root_company.id:
            selected_company = None

    if request.method == "POST":
        action = request.POST.get("action")
        company_id = (request.POST.get("company_id") or "").strip()
        payload = {
            "company_name": request.POST.get("company_name", "").strip(),
            "domain": request.POST.get("domain", "").strip(),
            "website": request.POST.get("website", "").strip(),
            "location": request.POST.get("location", "").strip(),
            "strength": request.POST.get("strength", "").strip(),
            "industry": request.POST.get("industry", "").strip(),
            "social_link": request.POST.get("social_link", "").strip(),
            "company_description": request.POST.get("company_description", "").strip(),
        }

        if action == "delete" and company_id:
            company = CompanyInfo.objects.filter(id=company_id).first()
            if company:
                if company.id == root_company.id:
                    messages.error(request, "Main company cannot be deleted from settings.")
                    return redirect("app_settings:company_info")
                if company.get_root_company().id != root_company.id:
                    messages.error(request, "Company not found.")
                    return redirect("app_settings:company_info")
                company.delete()
                messages.success(request, "Company deleted.")
            else:
                messages.error(request, "Company not found.")
            return redirect("app_settings:company_info")

        if action == "save":
            if not payload["company_name"]:
                messages.error(request, "Company Name is required.")
                return redirect("app_settings:company_info")

            if company_id:
                company = CompanyInfo.objects.filter(id=company_id).first()
                if not company:
                    messages.error(request, "Company not found.")
                    return redirect("app_settings:company_info")
                if company.get_root_company().id != root_company.id:
                    messages.error(request, "Company not found.")
                    return redirect("app_settings:company_info")
                for key, value in payload.items():
                    setattr(company, key, value)
                if request.FILES.get("logo"):
                    company.logo = request.FILES["logo"]
                company.save()
                messages.success(request, "Company information updated.")
            else:
                company = CompanyInfo(**payload)
                # Settings-side create = sub company under the current main company.
                company.parent_company = root_company
                if request.FILES.get("logo"):
                    company.logo = request.FILES["logo"]
                company.save()
                messages.success(request, "Company information saved.")
            return redirect("app_settings:company_info")

        if action == "preview":
            preview_data = payload
            preview_data["logo_name"] = request.FILES.get("logo").name if request.FILES.get("logo") else ""
            messages.info(request, "Preview generated below.")

        if action == "share":
            messages.success(request, "Share action triggered.")
            return redirect("app_settings:company_info")

    records = list(
        CompanyInfo.objects.filter(Q(id=root_company.id) | Q(parent_company=root_company)).order_by("company_name")
    )

    return render(
        request,
        "app_settings/company/info.html",
        {
            "settings_menu_active": "company",
            "records": records,
            "preview_data": preview_data,
            "selected_company": selected_company,
        },
    )


def company_careers_view(request):
    return render(
        request,
        "app_settings/company/careers.html",
        {"settings_menu_active": "company"},
    )


def role_view(request):
    if request.method == "POST":
        action = request.POST.get("action", "create_role").strip()
        role_id = request.POST.get("role_id", "").strip()
        if action == "delete_role" and role_id:
            role = RoleMaster.objects.filter(id=role_id).first()
            if role:
                role.delete()
                messages.success(request, "Role deleted successfully.")
            return redirect("app_settings:role")

        role_name = request.POST.get("role_name", "").strip()
        role_description = request.POST.get("role_description", "").strip()
        role_level = request.POST.get("role_level", "").strip()
        status = request.POST.get("status", "").strip()

        if not role_name or not role_level or not status:
            messages.error(request, "Role Name, Role Level, and Status are required.")
            return redirect("app_settings:role")

        try:
            if action == "update_role" and role_id:
                role = RoleMaster.objects.filter(id=role_id).first()
                if role:
                    role.role_name = role_name
                    role.role_description = role_description
                    role.role_level = role_level
                    role.status = status
                    role.save()
                    messages.success(request, "Role updated successfully.")
                else:
                    messages.error(request, "Role not found.")
            else:
                RoleMaster.objects.create(
                    role_name=role_name,
                    role_description=role_description,
                    role_level=role_level,
                    status=status,
                    created_by=request.session.get("login_user_name", "Administrator"),
                )
                messages.success(request, "Role created successfully.")
        except IntegrityError:
            messages.error(request, "Role Name already exists.")
        return redirect("app_settings:role")

    last_role = RoleMaster.objects.order_by("-id").first()
    next_id = (last_role.id + 1) if last_role else 1
    next_role_id = f"ROLE{next_id:04d}"

    records = list(RoleMaster.objects.all())
    edit_id = request.GET.get("edit", "").strip()
    selected_role = RoleMaster.objects.filter(id=edit_id).first() if edit_id else None

    return render(
        request,
        "app_settings/role.html",
        {
            "settings_menu_active": "role",
            "records": records,
            "next_role_id": next_role_id,
            "current_user_name": request.session.get("login_user_name", "Administrator"),
            "selected_role": selected_role,
        },
    )


def user_master_view(request):
    user_master_url = reverse("app_settings:user_master")
    if request.method == "POST":
        action = request.POST.get("action", "create_user")
        current_user_name = request.session.get("login_user_name", "Administrator")

        if action == "create_user":
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            email_id = request.POST.get("email_id", "").strip()
            mobile_number = request.POST.get("mobile_number", "").strip()
            username = request.POST.get("username", "").strip()
            password = request.POST.get("password", "").strip()
            confirm_password = request.POST.get("confirm_password", "").strip()
            department = request.POST.get("department", "").strip()
            designation = request.POST.get("designation", "").strip()
            role = request.POST.get("role", "").strip()
            team = request.POST.get("team", "").strip()
            reporting_manager = request.POST.get("reporting_manager", "").strip()
            location = request.POST.get("location", "").strip()
            status = request.POST.get("status", "").strip()
            date_of_joining = request.POST.get("date_of_joining", "").strip() or None
            allowed_modules = request.POST.getlist("allowed_modules")
            ip_restriction_enabled = request.POST.get("ip_restriction_enabled") == "on"
            two_factor_enabled = request.POST.get("two_factor_enabled") == "on"

            if not first_name or not last_name or not email_id or not mobile_number or not password or not confirm_password or not department or not role or not status or not allowed_modules:
                messages.error(request, "Please fill all mandatory fields.")
                return redirect("app_settings:user_master")

            if password != confirm_password:
                messages.error(request, "Password and Confirm Password must match.")
                return redirect("app_settings:user_master")

            try:
                user = UserMaster(
                    first_name=first_name,
                    last_name=last_name,
                    email_id=email_id,
                    mobile_number=mobile_number,
                    username=username,
                    password=password,
                    department=department,
                    designation=designation,
                    role=role,
                    team=team,
                    reporting_manager=reporting_manager,
                    location=location,
                    employee_code="",
                    status=status,
                    date_of_joining=date_of_joining,
                    allowed_modules=",".join(allowed_modules),
                    ip_restriction_enabled=ip_restriction_enabled,
                    two_factor_enabled=two_factor_enabled,
                    created_by=current_user_name,
                )
                if request.FILES.get("profile_photo"):
                    user.profile_photo = request.FILES["profile_photo"]
                user.save()
                UserMasterAudit.objects.create(
                    user=user,
                    action="User Created",
                    details=f"Created with role {user.role} and status {user.status}.",
                    changed_by=current_user_name,
                )
                messages.success(request, "User created successfully.")
            except IntegrityError:
                messages.error(request, "Email already exists.")
            return redirect("app_settings:user_master")

        user_id = request.POST.get("user_id")
        user = UserMaster.objects.filter(pk=user_id).first()
        if not user:
            messages.info(request, "Demo entry is view-only in UI mode.")
            return redirect(f"{user_master_url}?edit={user_id}")

        if action == "update_user":
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            email_id = request.POST.get("email_id", "").strip()
            mobile_number = request.POST.get("mobile_number", "").strip()
            department = request.POST.get("department", "").strip()
            designation = request.POST.get("designation", "").strip()
            role = request.POST.get("role", "").strip()
            status = request.POST.get("status", "").strip()
            team = request.POST.get("team", "").strip()
            reporting_manager = request.POST.get("reporting_manager", "").strip()
            location = request.POST.get("location", "").strip()
            allowed_modules = request.POST.getlist("allowed_modules")

            if not first_name or not last_name or not email_id or not mobile_number or not department or not role or not status or not allowed_modules:
                messages.error(request, "Please fill all mandatory fields in Update User form.")
                return redirect(f"{user_master_url}?edit={user.id}")

            user.first_name = first_name
            user.last_name = last_name
            user.email_id = email_id
            user.mobile_number = mobile_number
            user.department = department
            user.designation = designation
            user.role = role
            user.status = status
            user.team = team
            user.reporting_manager = reporting_manager
            user.location = location
            user.allowed_modules = ",".join(allowed_modules)
            try:
                user.save()
                UserMasterAudit.objects.create(
                    user=user,
                    action="User Updated",
                    details=f"Updated role={user.role}, status={user.status}, reporting_manager={user.reporting_manager or '-'}",
                    changed_by=current_user_name,
                )
                messages.success(request, "User updated successfully.")
            except IntegrityError:
                messages.error(request, "Email already exists.")
            return redirect(f"{user_master_url}?edit={user.id}")

        if action == "reset_password":
            user.password = "Temp@12345"
            user.save(update_fields=["password", "full_name"])
            UserMasterAudit.objects.create(
                user=user,
                action="Password Reset",
                details="Password reset to temporary password.",
                changed_by=current_user_name,
            )
            messages.success(request, f"Password reset for {user.full_name}. Temporary password: Temp@12345")
            return redirect(f"{user_master_url}?edit={user.id}")

        if action == "disable_login":
            user.status = "Suspended"
            user.save(update_fields=["status", "full_name"])
            UserMasterAudit.objects.create(
                user=user,
                action="Login Disabled",
                details="User status changed to Suspended.",
                changed_by=current_user_name,
            )
            messages.success(request, f"Login disabled for {user.full_name}.")
            return redirect(f"{user_master_url}?edit={user.id}")

        if action == "delete_user":
            deleted_name = user.full_name
            user.delete()
            messages.success(request, f"User deleted: {deleted_name}.")
            return redirect("app_settings:user_master")

        return redirect("app_settings:user_master")

    last_user = UserMaster.objects.order_by("-id").first()
    next_id = (last_user.id + 1) if last_user else 1
    next_user_id = f"USR{next_id:04d}"
    next_employee_code = f"EMP{next_id:04d}"
    selected_user = None
    selected_user_audits = []
    edit_id = request.GET.get("edit")
    if edit_id:
        selected_user = UserMaster.objects.filter(pk=edit_id).first()
        if selected_user:
            selected_user.selected_modules = [
                m.strip() for m in (selected_user.allowed_modules or "").split(",") if m.strip()
            ]
            selected_user_audits = selected_user.audit_logs.all()[:10]

    records = list(UserMaster.objects.all())

    role_options = list(RoleMaster.objects.values_list("role_name", flat=True))

    team_options = list(RecruitmentTeamMaster.objects.values_list("team_name", flat=True))
    if not team_options:
        team_options = ["IT Hiring Team", "QA Hiring Team"]

    manager_options = list(
        UserMaster.objects.exclude(full_name__exact="")
        .order_by("full_name", "email_id")
        .values("full_name", "mobile_number", "email_id")
    )

    context = {
        "settings_menu_active": "user_master",
        "next_user_id": next_user_id,
        "next_employee_code": next_employee_code,
        "current_user_name": request.session.get("login_user_name", "Administrator"),
        "role_options": role_options,
        "team_options": team_options,
        "manager_options": manager_options,
        "department_options": ["HR", "IT", "Finance", "Operations", "Admin"],
        "location_options": ["Head Office", "Branch Office", "Remote"],
        "module_options": USER_MODULE_OPTIONS,
        "records": records,
        "selected_user": selected_user,
        "selected_user_audits": selected_user_audits,
    }
    return render(request, "app_settings/user_master.html", context)


def notification_open_view(request, notification_id):
    login_user_name = (request.session.get("login_user_name") or "").strip()
    login_user_email = (request.session.get("login_user_email") or "").strip()
    notification = get_object_or_404(UiNotification, id=notification_id)
    recipient = notification.recipient_name.strip().lower()
    if not (login_user_name or login_user_email) or recipient not in {
        login_user_name.lower(),
        login_user_email.lower(),
    }:
        messages.error(request, "Notification access denied.")
        return redirect("/")
    target = (notification.link or "").strip()
    if (
        target.startswith("/interview-evaluation/panel/")
        or target.startswith("/interview-recording/panel/")
        or target.startswith("/recruitment-teams/")
    ):
        target = target.split("?", 1)[0]
    # Remove duplicates for the same recipient and payload.
    UiNotification.objects.filter(
        recipient_name__iexact=notification.recipient_name,
        title=notification.title,
        message=notification.message,
        link=notification.link,
    ).delete()
    if target:
        return redirect(target)
    return redirect(request.META.get("HTTP_REFERER", "/"))


def permission_management_view(request):
    if request.method == "POST":
        action = request.POST.get("action", "create_permission").strip()
        record_id = request.POST.get("record_id", "").strip()
        if action == "delete_permission" and record_id:
            record = RolePermissionSetup.objects.filter(id=record_id).first()
            if record:
                record.delete()
                messages.success(request, "Permission setup deleted.")
            return redirect("app_settings:permission_management")

        role = request.POST.get("role", "").strip()
        module = request.POST.get("module", "").strip()
        sub_module = request.POST.get("sub_module", "").strip()

        if not role or not module or not sub_module:
            messages.error(request, "Role, Module, and Sub Module are required.")
            return redirect("app_settings:permission_management")

        payload = {
            "role": role,
            "module": module,
            "sub_module": sub_module,
            "create_permission": request.POST.get("create_permission") == "on",
            "view_permission": request.POST.get("view_permission") == "on",
            "edit_permission": request.POST.get("edit_permission") == "on",
            "delete_permission": request.POST.get("delete_permission") == "on",
            "approve_permission": request.POST.get("approve_permission") == "on",
            "export_permission": request.POST.get("export_permission") == "on",
            "download_permission": request.POST.get("download_permission") == "on",
            "assign_permission": request.POST.get("assign_permission") == "on",
            "full_access": request.POST.get("full_access") == "on",
        }
        if action == "update_permission" and record_id:
            record = RolePermissionSetup.objects.filter(id=record_id).first()
            if record:
                for key, value in payload.items():
                    setattr(record, key, value)
                record.save()
                messages.success(request, "Permission setup updated.")
            else:
                messages.error(request, "Permission record not found.")
        else:
            RolePermissionSetup.objects.create(**payload)
            messages.success(request, "Permission setup saved.")
        return redirect("app_settings:permission_management")

    records = list(RolePermissionSetup.objects.all())
    edit_id = request.GET.get("edit", "").strip()
    selected_permission = RolePermissionSetup.objects.filter(id=edit_id).first() if edit_id else None

    role_options = list(RoleMaster.objects.values_list("role_name", flat=True))
    permission_module_options = [item["label"] for item in USER_MODULE_OPTIONS]
    selected_module_name = selected_permission.module if selected_permission else (permission_module_options[0] if permission_module_options else "")
    sub_module_options = PERMISSION_SUBMODULE_MAP.get(selected_module_name, ["General"])

    context = {
        "settings_menu_active": "permission_management",
        "role_options": role_options,
        "module_options": permission_module_options,
        "sub_module_options": sub_module_options,
        "sub_module_map": PERMISSION_SUBMODULE_MAP,
        "records": records,
        "selected_permission": selected_permission,
    }
    return render(request, "app_settings/permission_management.html", context)


def global_audit_logs_view(request):
    import json
    # pyrefly: ignore [missing-import]
    from .models import GlobalAuditLog, UserMasterAudit
    query = (request.GET.get("q") or "").strip()
    module_filter = (request.GET.get("module") or "").strip()
    
    logs = GlobalAuditLog.objects.select_related("candidate", "performed_by").order_by("-timestamp")
    
    if module_filter:
        logs = logs.filter(module=module_filter)
        
    if query:
        logs = logs.filter(
            Q(candidate__full_name__icontains=query) |
            Q(candidate__email__icontains=query) |
            Q(action__icontains=query) |
            Q(performed_by__full_name__icontains=query)
        )

    # Include legacy "Settings Activity" audit logs (UserMasterAudit) in the global view so
    # the Global Audit Logs page truly reflects create/update/delete admin actions.
    user_audits = UserMasterAudit.objects.select_related("user").order_by("-changed_at")
    if module_filter and module_filter != "app_settings":
        user_audits = user_audits.none()
    if query:
        user_audits = user_audits.filter(
            Q(user__full_name__icontains=query)
            | Q(user__email_id__icontains=query)
            | Q(action__icontains=query)
            | Q(changed_by__icontains=query)
            | Q(details__icontains=query)
        )

    rows = []
    for log in logs[:1000]:
        rows.append(
            {
                "id": f"g-{log.id}",
                "module": log.module,
                "module_display": log.get_module_display(),
                "action": log.action or "",
                "formatted_action": (log.action or "").replace("_", " ").title(),
                "performed_by_name": getattr(log.performed_by, "full_name", None) or "System",
                "performed_by_role": getattr(log.performed_by, "role", None) or "",
                "entity_name": getattr(log.candidate, "full_name", None) or "",
                "entity_sub": getattr(log.candidate, "email", None) or "",
                "timestamp": log.timestamp,
                "details_json": json.dumps(
                    {
                        "module": log.module,
                        "action": log.action,
                        "performed_by": getattr(log.performed_by, "full_name", None),
                        "candidate": getattr(log.candidate, "full_name", None),
                        "ip_address": log.ip_address,
                        "user_agent": log.user_agent,
                        "details": log.details or {},
                    },
                    ensure_ascii=False,
                ),
            }
        )

    for audit in user_audits[:1000]:
        rows.append(
            {
                "id": f"u-{audit.id}",
                "module": "app_settings",
                "module_display": "Settings",
                "action": audit.action or "",
                "formatted_action": (audit.action or "").replace("_", " ").title(),
                "performed_by_name": audit.changed_by or "System",
                "performed_by_role": "",
                "entity_name": getattr(audit.user, "full_name", None) or "",
                "entity_sub": getattr(audit.user, "email_id", None) or "",
                "timestamp": audit.changed_at,
                "details_json": json.dumps(
                    {
                        "module": "app_settings",
                        "action": audit.action,
                        "changed_by": audit.changed_by,
                        "target_user": getattr(audit.user, "user_id", None),
                        "details": audit.details or "",
                    },
                    ensure_ascii=False,
                ),
            }
        )

    rows.sort(key=lambda r: r["timestamp"], reverse=True)
    rows = rows[:1000]
    for row in rows:
        row["detail_url"] = reverse("app_settings:audit_log_entry", kwargs={"log_key": row["id"]})

    modules = GlobalAuditLog.MODULE_CHOICES

    return render(request, "app_settings/global_audit_logs.html", {
        "logs": rows,
        "query": query,
        "module_filter": module_filter,
        "modules": modules,
        "settings_menu_active": "global_audit_logs"
    })


def audit_log_entry_view(request, log_key):
    import json
    # pyrefly: ignore [missing-import]
    from .models import GlobalAuditLog, UserMasterAudit

    log_key = (log_key or "").strip()
    entry = None
    entry_type = None

    if log_key.startswith("g-"):
        entry_type = "global"
        pk = log_key.split("-", 1)[1]
        entry = get_object_or_404(
            GlobalAuditLog.objects.select_related("candidate", "performed_by"),
            pk=pk,
        )
        details_obj = {
            "module": entry.module,
            "action": entry.action,
            "performed_by": getattr(entry.performed_by, "full_name", None),
            "performed_by_role": getattr(entry.performed_by, "role", None),
            "candidate": getattr(entry.candidate, "full_name", None),
            "candidate_email": getattr(entry.candidate, "email", None),
            "ip_address": entry.ip_address,
            "user_agent": entry.user_agent,
            "details": entry.details or {},
            "timestamp": entry.timestamp.isoformat() if entry.timestamp else None,
        }
        title = "Audit Entry"
        subtitle = f"{entry.get_module_display()} • {(entry.action or '').replace('_', ' ').title()}"

    elif log_key.startswith("u-"):
        entry_type = "user_audit"
        pk = log_key.split("-", 1)[1]
        entry = get_object_or_404(
            UserMasterAudit.objects.select_related("user"),
            pk=pk,
        )
        details_obj = {
            "module": "app_settings",
            "action": entry.action,
            "changed_by": entry.changed_by,
            "target_user_id": getattr(entry.user, "user_id", None),
            "target_user_name": getattr(entry.user, "full_name", None),
            "target_user_email": getattr(entry.user, "email_id", None),
            "details": entry.details or "",
            "timestamp": entry.changed_at.isoformat() if entry.changed_at else None,
        }
        title = "Settings Activity"
        subtitle = f"Settings • {(entry.action or '').replace('_', ' ').title()}"

    else:
        raise Http404("Invalid audit entry key.")

    return render(
        request,
        "app_settings/audit_log_entry.html",
        {
            "settings_menu_active": "global_audit_logs",
            "log_key": log_key,
            "entry_type": entry_type,
            "title": title,
            "subtitle": subtitle,
            "details_pretty": json.dumps(details_obj, indent=2, ensure_ascii=False),
        },
    )

def onboarding_audit_logs_view(request):
    # pyrefly: ignore [missing-import]
    from .models import GlobalAuditLog

    query = (request.GET.get("q") or "").strip()
    logs = (
        GlobalAuditLog.objects.filter(module="onboarding")
        .select_related("candidate", "performed_by")
        .order_by("-timestamp")
    )

    if query:
        logs = logs.filter(
            Q(candidate__full_name__icontains=query)
            | Q(candidate__email__icontains=query)
            | Q(action__icontains=query)
            | Q(performed_by__full_name__icontains=query)
        )

    rows = list(logs[:500])
    for log in rows:
        log.formatted_action = (log.action or "").replace("_", " ").title()

    return render(
        request,
        "app_settings/onboarding_audit_logs.html",
        {"logs": rows, "query": query, "settings_menu_active": "onboarding_audit_logs"},
    )


def task_sla_settings_view(request):
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "save_sla":
            for priority in ["urgent", "high", "medium", "low"]:
                resp = int((request.POST.get(f"resp_{priority}") or "0").strip() or 0)
                res = int((request.POST.get(f"res_{priority}") or "0").strip() or 0)
                if resp <= 0 or res <= 0:
                    continue
                TaskSlaPrioritySetting.objects.update_or_create(
                    priority=priority,
                    defaults={"first_response_hours": resp, "resolution_hours": res},
                )
            messages.success(request, "Task SLA settings saved.")
            return redirect("app_settings:task_sla_settings")

    rows = list(TaskSlaPrioritySetting.objects.all())
    by_key = {r.priority: r for r in rows}
    ordered = []
    for key, label in TaskSlaPrioritySetting.PRIORITY_CHOICES:
        ordered.append(
            {
                "key": key,
                "label": label,
                "resp": getattr(by_key.get(key), "first_response_hours", 0) or 0,
                "res": getattr(by_key.get(key), "resolution_hours", 0) or 0,
            }
        )

    return render(
        request,
        "app_settings/task_sla_settings.html",
        {
            "settings_menu_active": "task_sla_settings",
            "rows": ordered,
        },
    )
